"""
cdd_sync_itn.py — ITN/LLIN CDD sync tracking (chad-user-sync-index-v1)

Separate module from cdd_sync.py by design (see analyze_itn.py's docstring for the
full "no touching original files" reasoning). Even though cdd_sync.py already has a
tenant-branching precedent (a `tenant == "ch"` special case), the instruction this
session has been separate files only — so this is additive, cdd_sync.py is untouched.

Confirmed this session (tenant=chad, campaign CMP-2026-06-03-000312):
  - chad-user-sync-index-v1 exists and Data.syncedUserId.keyword / Data.taskDates
    are valid, correct fields (same names cdd_sync.py already uses for other
    non-admin-console tenants).
  - Data.role.keyword IS present directly on sync records (unlike having to
    cross-reference the staff index) — confirmed only two roles ever sync:
    DISTRIBUTOR_REGISTRAR (the real CDD role for chad) and WAREHOUSE_MANAGER.
  - "DISTRIBUTOR" (the role cdd_sync.py's existing tenants filter by) is NOT the
    right role for chad — DISTRIBUTOR_REGISTRAR is ~96% of the field-staff
    population; plain DISTRIBUTOR is a small minority (540 of 14,857+).

KNOWN GAP, deliberately not papered over: this module can report how many CDDs
are ACTIVELY SYNCING (a real, verified number), but NOT a roster-based coverage
percentage ("X of Y total registered CDDs"). Getting a correct denominator
requires scoping the STAFF index (chad-project-staff-index-v1) to just this
campaign's project hierarchy — proven this session that neither
Data.additionalDetails.cycleIndex (cycle "03" doesn't exist on ANY staff record,
staff assignments aren't re-tagged per cycle) nor Data.projectTypeId alone
(spans every historical LLIN campaign/pilot for this tenant) can scope it
correctly. The real fix needs a DB query (chad.project, projecthierarchy LIKE
'<campaign root>.%') to get every projectId under this specific campaign, then
filter staff by Data.projectId IN (...) — not yet built, deliberately left as a
TODO rather than shipping a number computed from an unverified/wrong roster.
"""
import logging
import os
from datetime import datetime, timedelta, timezone

import pandas as pd
import requests
import urllib3
from openpyxl import Workbook

urllib3.disable_warnings()

# Shares only the generic Excel-styling primitives from pipeline.core. NOT reusing
# cdd_sync.py's _write_summary/_write_df/_build_rows/_sync_status — those are built
# around a fixed "Day 1..Day N of a short campaign" model, which doesn't fit a
# continuously-running, months-long ITN campaign (see _roster_status below).
from pipeline.core.excel import (
    SYNC_HDR_FILL, SYNC_LOW_FILL, SYNC_NEVER_FILL, SYNC_TOTAL_FILL, style_sync_cell,
)

log = logging.getLogger(__name__)

# Confirmed the real field-CDD role for chad (~96% of field staff; plain
# DISTRIBUTOR is a small minority). Overridable per deployment via CDD_ROLE_ITN
# so a new ITN tenant with a different role needs no code edit. Separate from
# cdd_sync.py's CDD_ROLE because one deployment runs SMC and ITN together.
DEFAULT_CDD_ROLE_ITN = "DISTRIBUTOR_REGISTRAR"


def _cdd_role():
    return (os.getenv("CDD_ROLE_ITN", "").strip() or DEFAULT_CDD_ROLE_ITN)

# Per-day Y/N matrix width cap (SMC template on a months-long campaign). The most
# recent MAX_DAY_COLS elapsed days get a column; older days stay counted in
# "Days Synced" and listed in "Sync Dates".
MAX_DAY_COLS = 31

# CONFIRMED this session: chad-user-sync-index-v1 DOES carry
# Data.additionalDetails.projectReferenceId, same as the task index — proven by
# a real sample record showing "CMP-2026-01-24-000223" (LLIN_phase1_tchad2026-Final),
# a DIFFERENT campaign than ours — i.e. this index genuinely mixes multiple
# campaigns' sync records together, confirming the scoping filter below is
# necessary, not just precautionary.


def _campaign_filter(cfg):
    """Same scoping field as analyze_itn.py's task-index filter, confirmed
    present on the sync index too."""
    if not cfg.get("campaign_number"):
        raise ValueError("campaign_number is required for ITN CDD sync reporting")
    return {"term": {"Data.additionalDetails.projectReferenceId.keyword": cfg["campaign_number"]}}


def _distinct_cdds_synced(cfg, date_str=None):
    """
    Distinct DISTRIBUTOR_REGISTRAR users with a sync record for THIS campaign
    (scoped via _campaign_filter — confirmed necessary, this index genuinely
    mixes multiple campaigns' records together), optionally scoped to one date
    (taskDates). No date filter = cumulative to date.
    Returns (count, doc_count) — doc_count included since it's a distinct signal
    from distinct-user count (many sync docs can belong to the same user/day).
    """
    filters = [_campaign_filter(cfg), {"term": {"Data.role.keyword": _cdd_role()}}]
    if date_str:
        filters.append({"term": {"Data.taskDates": date_str}})

    body = {
        "size": 0,
        "query": {"bool": {"filter": filters}},
        "aggs": {"distinct_cdds": {"cardinality": {"field": "Data.syncedUserId.keyword"}}},
    }
    r = requests.post(
        f"{cfg['es_url']}/{cfg['ES_INDEX_SYNC']}/_search",
        json=body, auth=cfg["es_auth"], verify=False, timeout=60,
    )
    r.raise_for_status()
    data = r.json()
    return data["aggregations"]["distinct_cdds"]["value"], data["hits"]["total"]["value"]


# ── time-cutoff sync count ────────────────────────────────────────────────────
# Mirrors cdd_sync.py's _count_synced_by_cutoff exactly (same createdTime-epoch-ms
# approach, confirmed present on chad's sync docs this session — sample doc showed
# Data.createdTime: 1785564126353). Only difference from the SPAQ version: campaign
# scoping field (additionalDetails.projectReferenceId, not campaignNumber) and role
# (DISTRIBUTOR_REGISTRAR, not DISTRIBUTOR) — both already established elsewhere in
# this file. Verified live: 418 distinct CDDs synced by 17:00 UTC today (9,493 sync
# docs before cutoff, out of >=10,000 total today's docs for this campaign+role).

def _count_synced_by_cutoff(cfg, cutoff_hour, cutoff_min=0):
    """
    Count unique CDDs (this campaign, DISTRIBUTOR_REGISTRAR) who synced TODAY
    before cutoff_hour:cutoff_min UTC. Uses Data.createdTime (epoch ms, server
    receipt timestamp). Returns int count, or None if query fails.
    """
    today = cfg["extract_date"].isoformat()
    cutoff_dt = datetime.strptime(
        f"{today}T{cutoff_hour:02d}:{cutoff_min:02d}:00", "%Y-%m-%dT%H:%M:%S"
    )
    cutoff_ms = int(cutoff_dt.replace(tzinfo=timezone.utc).timestamp() * 1000)

    filters = [
        _campaign_filter(cfg),
        {"term": {"Data.role.keyword": _cdd_role()}},
        {"term": {"Data.taskDates": today}},
        {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
    ]
    body = {
        "size": 0,
        "query": {"bool": {"filter": filters}},
        "aggs": {"unique_synced": {"cardinality": {"field": "Data.syncedUserId.keyword"}}},
    }
    try:
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_SYNC']}/_search",
            json=body, auth=cfg["es_auth"], verify=False, timeout=30,
        )
        r.raise_for_status()
        count = r.json()["aggregations"]["unique_synced"]["value"]
        log.info(f"[cdd_sync_itn] synced by {cutoff_hour:02d}:{cutoff_min:02d} UTC: {count:,}")
        return count
    except Exception as e:
        log.warning(f"[cdd_sync_itn] time-cutoff query failed (non-fatal): {e}")
        return None


def _get_synced_keys_by_cutoff(cfg, cutoff_hour, cutoff_min=0):
    """
    Lowercased syncedUserId set of CDDs who synced TODAY before the cutoff —
    mirrors cdd_sync.py's _get_synced_keys_by_cutoff (drives the SMC-template
    "NOT SYNCED BY 1730" tab). Returns None on failure so the tab is skipped
    rather than written empty-but-wrong.
    """
    today = cfg["extract_date"].isoformat()
    cutoff_dt = datetime.strptime(
        f"{today}T{cutoff_hour:02d}:{cutoff_min:02d}:00", "%Y-%m-%dT%H:%M:%S"
    )
    cutoff_ms = int(cutoff_dt.replace(tzinfo=timezone.utc).timestamp() * 1000)

    filters = [
        _campaign_filter(cfg),
        {"term": {"Data.role.keyword": _cdd_role()}},
        {"term": {"Data.taskDates": today}},
        {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
    ]
    keys = set()
    after = None
    try:
        while True:
            composite = {"size": 1000,
                         "sources": [{"uid": {"terms": {"field": "Data.syncedUserId.keyword"}}}]}
            if after:
                composite["after"] = after
            body = {"size": 0, "query": {"bool": {"filter": filters}},
                    "aggs": {"by_user": {"composite": composite}}}
            r = requests.post(
                f"{cfg['es_url']}/{cfg['ES_INDEX_SYNC']}/_search",
                json=body, auth=cfg["es_auth"], verify=False, timeout=60,
            )
            r.raise_for_status()
            agg = r.json()["aggregations"]["by_user"]
            keys.update(b["key"]["uid"].lower() for b in agg["buckets"])
            after = agg.get("after_key")
            if not after:
                break
        return keys
    except Exception as e:
        log.warning(f"[cdd_sync_itn] cutoff key query failed (non-fatal): {e}")
        return None


# ── CDD roster (sync-derived) ────────────────────────────────────
# Derived FROM sync records rather than the staff index (no correct
# campaign-scoping field on chad-project-staff-index-v1 - see module docstring).
# Consequence: a CDD who never synced cannot appear, so NEVER SYNCED always reads
# 0. Structural, not a bug; disclosed again in the SUMMARY tab. The column is kept
# so the table shape matches cdd_sync.py.
# Thresholds match cdd_sync._sync_status(n, day): n = distinct days synced,
# day = elapsed campaign day (not cfg["DAY"], which defaults to 4).

def _fetch_cdd_roster(cfg):
    """
    Composite aggregation over syncedUserId, cumulative (no date filter) —
    scales to however long this campaign has been running without scrolling every
    individual sync doc (this campaign alone has >=10,000 sync docs for a single
    day; scrolling the full cumulative history would be wasteful when all we need
    is one row per distinct CDD). Sub-aggs per user: distinct days synced, total
    sync record count (bucket doc_count), first/last sync time, and a top_hits
    sample of the most recent record's syncedUserName + boundaryHierarchy
    (province/district/facility) — confirmed present on chad's sync docs this
    session (sample doc carried province=OUADDAI, district=ADRE, sppSfd=CS KATARFA).
    """
    filters = [_campaign_filter(cfg), {"term": {"Data.role.keyword": _cdd_role()}}]
    rows = {}
    after = None
    while True:
        composite = {"size": 1000, "sources": [{"uid": {"terms": {"field": "Data.syncedUserId.keyword"}}}]}
        if after:
            composite["after"] = after
        body = {
            "size": 0,
            "query": {"bool": {"filter": filters}},
            "aggs": {
                "by_user": {
                    "composite": composite,
                    "aggs": {
                        "distinct_days": {"cardinality": {"field": "Data.taskDates"}},
                        # actual per-user sync-date set — needed for the SMC-template
                        # per-day Y/N matrix + "Sync Dates" column. format required:
                        # taskDates is date-mapped, without it keys come back as epoch ms
                        "sync_days": {"terms": {"field": "Data.taskDates", "size": 400,
                                                "format": "yyyy-MM-dd"}},
                        "last_sync_ms":  {"max": {"field": "Data.createdTime"}},
                        "first_sync_ms": {"min": {"field": "Data.createdTime"}},
                        "latest": {
                            "top_hits": {
                                "size": 1,
                                "sort": [{"Data.createdTime": "desc"}],
                                "_source": ["Data.syncedUserName", "Data.boundaryHierarchy"],
                            }
                        },
                    },
                }
            },
        }
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_SYNC']}/_search",
            json=body, auth=cfg["es_auth"], verify=False, timeout=120,
        )
        r.raise_for_status()
        data = r.json()
        buckets = data["aggregations"]["by_user"]["buckets"]
        for b in buckets:
            uid = b["key"]["uid"]
            hits = b["latest"]["hits"]["hits"]
            src = hits[0]["_source"]["Data"] if hits else {}
            bh = src.get("boundaryHierarchy") or {}
            last_ms = b["last_sync_ms"]["value"]
            first_ms = b["first_sync_ms"]["value"]
            rows[uid] = {
                "user_id": uid,
                "username": src.get("syncedUserName", ""),
                "province": bh.get("province", ""),
                "district": bh.get("district", ""),
                "facility": bh.get("sppSfd", ""),
                "records": b["doc_count"],
                "distinct_days": int(b["distinct_days"]["value"]),
                "dates": {d["key_as_string"] for d in b["sync_days"]["buckets"]},
                "first_sync": (
                    datetime.fromtimestamp(first_ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
                    if first_ms else ""
                ),
                "last_sync": (
                    datetime.fromtimestamp(last_ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
                    if last_ms else ""
                ),
            }
        after = data["aggregations"]["by_user"].get("after_key")
        if not after:
            break
    log.info(f"[cdd_sync_itn] roster (sync-derived): {len(rows):,} distinct CDDs")
    return rows


def _roster_status(n_days, elapsed_day):
    """Identical thresholds to cdd_sync.py's _sync_status(n, day)."""
    if n_days == elapsed_day: return "HIGH"
    elif n_days >= 3:         return "MODERATE"
    elif n_days >= 1:         return "LOW"
    else:                     return "NEVER SYNCED"


def _build_roster_rows(cfg):
    roster = _fetch_cdd_roster(cfg)
    extract_date   = cfg.get("extract_date")
    campaign_start = cfg.get("campaign_start")
    elapsed_day = (extract_date - campaign_start).days + 1 if (extract_date and campaign_start) else None

    # SMC-template day columns (Day 1..N Y/N matrix, same labels as cdd_sync.py's
    # _build_rows). SMC campaigns are 4-5 days so the matrix spans CAMPAIGN_DATES
    # whole; an ITN campaign runs for months, so the matrix covers elapsed days
    # only (campaign_start..extract_date, clipped to campaign_end) and is capped
    # at the most recent MAX_DAY_COLS days — older days are still counted in
    # "Days Synced" and listed in "Sync Dates", never silently dropped.
    day_labels = {}
    window_dates = []
    if campaign_start and extract_date:
        end = extract_date
        if cfg.get("campaign_end"):
            end = min(end, cfg["campaign_end"])
        n_days = (end - campaign_start).days + 1
        window_dates = [(campaign_start + timedelta(days=i)).isoformat() for i in range(n_days)]
        offset = max(0, n_days - MAX_DAY_COLS)
        if offset:
            log.info(f"[cdd_sync_itn] day matrix shows Day {offset+1}-{n_days} "
                     f"(most recent {MAX_DAY_COLS} of {n_days} elapsed days); "
                     f"earlier days remain in Days Synced / Sync Dates")
        for i in range(offset, n_days):
            d = campaign_start + timedelta(days=i)
            day_labels[d.isoformat()] = f"Day {i+1} ({d.day} {d.strftime('%b')})"
    window_set = set(window_dates)

    all_rows = []
    for uid, info in roster.items():
        # scope to the campaign window, same as cdd_sync.py (its sync agg filters
        # to CAMPAIGN_DATES) — pre-campaign syncs (e.g. training/test days) no
        # longer inflate Days Synced or Status
        dates_in_window = info["dates"] & window_set if window_set else info["dates"]
        n_days_synced = len(dates_in_window)
        status = _roster_status(n_days_synced, elapsed_day) if elapsed_day else "LOW"
        rec = {
            "District":        info["district"] or info["province"] or "Unknown",
            "Health Facility": info["facility"],
            "Username":        info["username"],
            "User ID":         info["user_id"],
            "Days Synced":     n_days_synced,
        }
        for dt, col in day_labels.items():
            rec[col] = "Y" if dt in dates_in_window else "N"
        rec["Status"]     = status
        rec["Sync Dates"] = ", ".join(sorted(dates_in_window))
        all_rows.append(rec)
    return all_rows, day_labels


# ── Excel writer ───────────────────────────────────────────────────────────────

def _write_df_itn(ws, df):
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        style_sync_cell(cell, fill=SYNC_HDR_FILL, bold=True, color="FFFFFF")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        status = getattr(row, "Status", None)
        row_fill = (
            SYNC_NEVER_FILL if status == "NEVER SYNCED"
            else SYNC_LOW_FILL if status == "LOW"
            else None
        )
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            style_sync_cell(cell, fill=row_fill)
    for ci, col in enumerate(df.columns, 1):
        lengths = [len(str(col))] + [len(str(v)) for v in df.iloc[:, ci - 1] if v is not None]
        max_len = max(lengths) if lengths else 10
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 30)


def _write_summary_itn(ws, all_rows):
    """Same column set/layout as cdd_sync.py's _write_summary: # | District |
    Total CDDs | HIGH | MODERATE | LOW | NEVER SYNCED | % Never Synced."""
    from collections import defaultdict
    dist_stats = defaultdict(lambda: {"total": 0, "HIGH": 0, "MODERATE": 0, "LOW": 0, "NEVER SYNCED": 0})
    for rec in all_rows:
        d = rec["District"] or "Unknown"
        dist_stats[d]["total"] += 1
        dist_stats[d][rec["Status"]] += 1

    cols = ["#", "District", "Total CDDs", "HIGH", "MODERATE", "LOW", "NEVER SYNCED", "% Never Synced"]
    for ci, h in enumerate(cols, 1):
        style_sync_cell(ws.cell(row=1, column=ci, value=h), fill=SYNC_HDR_FILL, bold=True, color="FFFFFF")

    for ri, (dist, s) in enumerate(sorted(dist_stats.items()), 2):
        pct = f"{s['NEVER SYNCED']/s['total']*100:.1f}%" if s["total"] else "-"
        vals = [ri - 1, dist, s["total"], s["HIGH"], s["MODERATE"], s["LOW"], s["NEVER SYNCED"], pct]
        for ci, val in enumerate(vals, 1):
            style_sync_cell(ws.cell(row=ri, column=ci, value=val))

    tr = len(dist_stats) + 2
    totals = [
        "", "GRAND TOTAL",
        sum(s["total"] for s in dist_stats.values()),
        sum(s["HIGH"] for s in dist_stats.values()),
        sum(s["MODERATE"] for s in dist_stats.values()),
        sum(s["LOW"] for s in dist_stats.values()),
        sum(s["NEVER SYNCED"] for s in dist_stats.values()),
        "",
    ]
    for ci, val in enumerate(totals, 1):
        style_sync_cell(ws.cell(row=tr, column=ci, value=val), fill=SYNC_TOTAL_FILL, bold=True)

    note_row = tr + 2
    ws.cell(note_row, 1,
            "Note: roster is derived from CDDs who have synced at least once for this "
            "campaign (no assigned-staff roster available). NEVER SYNCED here means no "
            "sync within the campaign window (campaign_start onward) — a CDD assigned "
            "but with zero syncs ever cannot appear in this workbook at all.")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(cols))
    ws.cell(note_row, 1).alignment = ws.cell(note_row, 1).alignment.copy(horizontal="left", wrap_text=True)

    return dist_stats, tr


def run(cfg):
    """
    Returns a dict of what's verified/available. `roster_total` and `coverage_pct`
    are deliberately None — see module docstring for why that denominator isn't
    solved yet. Callers (report_itn.py) must handle None gracefully, not assume
    a coverage percentage exists.
    """
    log.info(f"[cdd_sync_itn] {cfg['state_name']} — checking CDD sync activity ...")

    today_str = cfg.get("extract_date").isoformat() if cfg.get("extract_date") else None

    cumulative_cdds, cumulative_docs = _distinct_cdds_synced(cfg)
    today_cdds, today_docs = (None, None)
    if today_str:
        today_cdds, today_docs = _distinct_cdds_synced(cfg, today_str)

    log.info(
        f"[cdd_sync_itn] cumulative distinct CDDs synced (this campaign): {cumulative_cdds:,} "
        f"({cumulative_docs:,} sync records)"
        + (f"; today: {today_cdds:,} ({today_docs:,} records)" if today_str else "")
    )

    # Time-cutoff sync counts. Same-day operational metric, meaningless once the
    # report spans the whole multi-month campaign — skip for cumulative runs, same
    # gating as cdd_sync.py's _write_summary caller.
    synced_by_1700 = synced_by_1730 = None
    if today_str and not cfg.get("cumulative"):
        synced_by_1700 = _count_synced_by_cutoff(cfg, 17, 0)
        synced_by_1730 = _count_synced_by_cutoff(cfg, 17, 30)

    # ── Excel workbook (sync-derived roster, SMC template) ──────────────────
    all_rows, day_labels = _build_roster_rows(cfg)
    out = None
    low_count = 0
    if all_rows:
        wb = Workbook()
        wb.remove(wb.active)

        ws_sum = wb.create_sheet("SUMMARY")
        dist_stats, last_row = _write_summary_itn(ws_sum, all_rows)

        # Append time-based stats below the note row, same layout as cdd_sync.py
        total_cdds = len(all_rows)
        time_rows = []
        if synced_by_1700 is not None:
            pct = f"{synced_by_1700/total_cdds*100:.1f}%" if total_cdds else "-"
            time_rows.append(("Synced by 17:00 today (UTC)", synced_by_1700, pct))
        if synced_by_1730 is not None:
            pct = f"{synced_by_1730/total_cdds*100:.1f}%" if total_cdds else "-"
            time_rows.append(("Synced by 17:30 today (UTC)", synced_by_1730, pct))
        for i, (label, count, pct) in enumerate(time_rows):
            r = last_row + 4 + i
            ws_sum.cell(r, 1, label)
            ws_sum.cell(r, 2, count)
            ws_sum.cell(r, 3, pct)
            for ci in range(1, 4):
                style_sync_cell(ws_sum.cell(r, ci), fill=SYNC_TOTAL_FILL, bold=True)

        # SMC-template columns: same shape as cdd_sync.py's run() (boundary col is
        # District here, LGA there — position and everything else identical)
        COLS = (["District", "Health Facility", "Username", "User ID", "Days Synced"]
                + list(day_labels.values()) + ["Status", "Sync Dates"])
        df_all = pd.DataFrame(all_rows).sort_values(["District", "Days Synced"])
        for col in COLS:
            if col not in df_all.columns:
                df_all[col] = ""

        # Per-district tabs (SMC: per-LGA tabs, full column set incl. day matrix)
        for dist in sorted(df_all["District"].unique()):
            df_d = df_all[df_all["District"] == dist][COLS].copy().reset_index(drop=True)
            df_d.insert(0, "#", range(1, len(df_d) + 1))
            safe = str(dist).translate(str.maketrans("/\\*?[]:", "-------"))[:31] or "Unknown"
            ws = wb.create_sheet(safe)
            _write_df_itn(ws, df_d)

        # NEVER SYNCED — SMC column set. Non-empty here only for CDDs whose syncs
        # all predate campaign_start (assigned-but-zero-syncs still can't appear —
        # roster is sync-derived, see SUMMARY note)
        df_never = df_all[df_all["Status"] == "NEVER SYNCED"][
            ["District", "Health Facility", "Username", "User ID"]
        ].reset_index(drop=True)
        df_never.insert(0, "#", range(1, len(df_never) + 1))
        ws_never = wb.create_sheet("NEVER SYNCED")
        _write_df_itn(ws_never, df_never)

        # LOW SYNCED — SMC column set
        df_low = df_all[df_all["Status"] == "LOW"][
            ["District", "Health Facility", "Username", "User ID",
             "Days Synced", "Sync Dates", "Status"]
        ].reset_index(drop=True)
        df_low.insert(0, "#", range(1, len(df_low) + 1))
        low_count = len(df_low)
        ws_low = wb.create_sheet("LOW SYNCED")
        _write_df_itn(ws_low, df_low)

        # NOT SYNCED BY 1730 — SMC tab (was NOT SYNCED TODAY): roster CDDs with no
        # sync record before 17:30 UTC today. Daily runs only.
        if today_str and not cfg.get("cumulative"):
            synced_keys_1730 = _get_synced_keys_by_cutoff(cfg, 17, 30)
            if synced_keys_1730 is not None:
                df_ns = df_all[~df_all["User ID"].str.lower().isin(synced_keys_1730)][
                    ["District", "Health Facility", "Username", "User ID"]
                ].reset_index(drop=True)
                df_ns.insert(0, "#", range(1, len(df_ns) + 1))
                ws_ns = wb.create_sheet("NOT SYNCED BY 1730")
                _write_df_itn(ws_ns, df_ns)
                log.info(f"[cdd_sync_itn] not synced by 17:30 UTC: {len(df_ns):,} CDDs")

        out = cfg["sync_xlsx"]
        wb.save(out)
        log.info(
            f"[cdd_sync_itn] saved -> {out}  "
            f"({total_cdds} CDDs in roster, {low_count} LOW)"
        )

    return {
        "role": _cdd_role(),
        "campaign_scoped": True,   # confirmed via additionalDetails.projectReferenceId
        "cumulative_cdds_synced": cumulative_cdds,
        "cumulative_sync_records": cumulative_docs,
        "today_cdds_synced": today_cdds,
        "today_sync_records": today_docs,
        "synced_by_1700": synced_by_1700,
        "synced_by_1730": synced_by_1730,
        "roster_total": None,     # TODO: needs DB projecthierarchy join — see docstring
        "coverage_pct": None,     # cannot compute without roster_total
        "roster_ever_synced": len(all_rows),
        "roster_low": low_count,
        "sync_xlsx": out,
    }
