"""
analyze.py — ES task index scroll + individual batch lookup → performance Excel
"""
import logging
import os
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from pipeline.core.checkpoint import load_checkpoint, save_checkpoint
from pipeline.core.es import scroll_batches
from pipeline.core.excel import (
    BANNER_FILL, FLAG_COLOR, HDR_FILL, TOTAL_FILL, WHITE_FILL, style_cell,
)

log = logging.getLogger(__name__)

# Name resolution runs in parallel batches against the individual, project-
# beneficiary and household-member indices. A failed batch used to be logged at
# WARNING and discarded: the run then reported SUCCESS while some children had no
# resolved name, so the "Missing Child Name" and "Missing HH Name" DQ columns
# silently overstated a data problem that was really a partial fetch failure.
# That is the shape of the 2026-07-02 bug where Missing HH Name read 103%.
#
# campaign_runner clears this before a run and inspects it after the analyze
# stage, so a partial fetch marks the run degraded instead of passing quietly.
NAME_BATCH_FAILURES = []

# Operator-facing reasons THIS run's numbers cannot be trusted, for conditions
# the pipeline deliberately survives rather than crashing on. campaign_runner
# clears this before a run and reads it after the analyze stage, turning each
# one into a "degraded" outcome that reaches the Run Log and Slack.
#
# Before this existed, both conditions below published a complete, plausible,
# green report: a wrong index prefix or campaign_number matched zero documents,
# and a missing target book made every coverage figure zero. The code logged
# "the report should not be shared" and then shared it.
RUN_DEGRADATIONS = []


TARGETS_ZERO = ("targets are ALL ZERO, so every coverage percentage and every facility status band in this report is meaningless. Cause: ")


def _degrade_run(reason):
    RUN_DEGRADATIONS.append(reason)
    log.error(f"[analyze] RUN DEGRADED - {reason}")



def _record_batch_failure(kind, exc):
    NAME_BATCH_FAILURES.append(f"{kind}: {type(exc).__name__}: {exc}")
    log.error(f"  {kind} batch FAILED — names from this batch will be missing, "
              f"which INFLATES the missing-name DQ columns: {exc}", exc_info=True)

_BATCH = 5000
_WORKERS = 8


def _build_campaign_filters(cfg):
    """
    Build ES filter clauses that identify this specific campaign in the task index.

    Only applied when task_campaign_filter=TRUE (opt-in).
    - Nigeria SMC states: FALSE — date range alone isolates the campaign.
    - AZM/non-admin (multiple project types share same tenant+date): TRUE.

    Priority (first match wins):
      1. is_admin_console=TRUE  → campaignNumber  (Nigeria admin, Chad admin)
      2. project_type_id set    → projectTypeId   (Togo non-admin)
      3. project_type set       → projectType     (AZM Nigeria/Congo)
      Appends cycleIndex filter if cycle_index is set (regardless of above).
    """
    if not cfg.get("task_campaign_filter", False):
        return []

    filters = []
    if cfg.get("is_admin_console") and cfg.get("campaign_number"):
        filters.append({"term": {"Data.campaignNumber.keyword": cfg["campaign_number"]}})
    elif cfg.get("project_type_id"):
        filters.append({"term": {"Data.projectTypeId.keyword": cfg["project_type_id"]}})
    elif cfg.get("project_type"):
        filters.append({"term": {"Data.projectType.keyword": cfg["project_type"]}})

    if cfg.get("cycle_index"):
        filters.append({"term": {"Data.additionalDetails.cycleIndex.keyword": cfg["cycle_index"]}})

    return filters



def _fetch_individual_names(cfg, ind_ids):
    """Batch-fetch child names from individual index. Returns {clientReferenceId: name}."""
    if not ind_ids:
        return {}
    batches = [ind_ids[i:i + _BATCH] for i in range(0, len(ind_ids), _BATCH)]
    name_map = {}

    def _one_batch(batch):
        q = {
            "size": _BATCH,
            "query": {"terms": {"clientReferenceId.keyword": batch}},
            "_source": ["clientReferenceId", "name"],
        }
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_IND']}/_search",
            json=q, auth=cfg["es_auth"], verify=False, timeout=60,
        )
        r.raise_for_status()
        result = {}
        for h in r.json()["hits"]["hits"]:
            src = h["_source"]
            cid = src.get("clientReferenceId", "")
            n   = src.get("name", {})
            given  = n.get("givenName")  or ""
            family = n.get("familyName") or ""
            result[cid] = f"{given} {family}".strip()
        return result

    log.info(f"  individual lookup: {len(ind_ids):,} IDs in {len(batches)} batches ...")
    with ThreadPoolExecutor(max_workers=_WORKERS) as ex:
        futures = {ex.submit(_one_batch, b): b for b in batches}
        for f in as_completed(futures):
            try:
                name_map.update(f.result())
            except Exception as e:
                _record_batch_failure("individual", e)
    log.info(f"  individual lookup: {len(name_map):,} names resolved")
    return name_map


def _map_beneficiary_refs_to_individual_ids(cfg, pb_refs):
    """Chad: project-beneficiary clientReferenceId -> the individual's clientReferenceId.
    The PB index is FLAT (no Data. wrapper); for an INDIVIDUAL beneficiary,
    beneficiaryClientReferenceId IS the individual's clientReferenceId.
    Returns {pb_clref_id: individual_clref_id}.
    """
    if not pb_refs:
        return {}
    batches = [pb_refs[i:i + _BATCH] for i in range(0, len(pb_refs), _BATCH)]
    out = {}

    def _one_batch(batch):
        q = {
            "size": _BATCH,
            "query": {"terms": {"clientReferenceId.keyword": batch}},
            "_source": ["clientReferenceId", "beneficiaryClientReferenceId"],
        }
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_PB']}/_search",
            json=q, auth=cfg["es_auth"], verify=False, timeout=60,
        )
        r.raise_for_status()
        res = {}
        for h in r.json()["hits"]["hits"]:
            s   = h["_source"]
            pb  = s.get("clientReferenceId", "")
            ind = s.get("beneficiaryClientReferenceId", "")
            if pb and ind:
                res[pb] = ind
        return res

    log.info(f"  PB->individual: {len(pb_refs):,} refs in {len(batches)} batches ...")
    with ThreadPoolExecutor(max_workers=_WORKERS) as ex:
        futures = {ex.submit(_one_batch, b): b for b in batches}
        for f in as_completed(futures):
            try:
                out.update(f.result())
            except Exception as e:
                _record_batch_failure("PB", e)
    log.info(f"  PB->individual: {len(out):,} beneficiaries resolved")
    return out


def _map_individual_ids_to_household_ids(cfg, ind_ids):
    """
    Step 4 (Togo pattern): for each individual ID find their household clientReferenceId.
    Queries household-member-index by Data.householdMember.individualClientReferenceId.
    Returns {ind_clref_id: hh_clref_id}.
    """
    if not ind_ids:
        return {}
    batches    = [ind_ids[i:i + _BATCH] for i in range(0, len(ind_ids), _BATCH)]
    member_map = {}

    def _one_batch(batch):
        q = {
            "size": _BATCH,
            "query": {"terms": {
                "Data.householdMember.individualClientReferenceId.keyword": batch
            }},
            "_source": [
                "Data.householdMember.individualClientReferenceId",
                "Data.householdMember.householdClientReferenceId",
            ],
        }
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_HH_MEMBER']}/_search",
            json=q, auth=cfg["es_auth"], verify=False, timeout=60,
        )
        r.raise_for_status()
        result = {}
        for h in r.json()["hits"]["hits"]:
            src    = h["_source"].get("Data", {}).get("householdMember", {})
            ind_id = src.get("individualClientReferenceId", "")
            hh_id  = src.get("householdClientReferenceId", "")
            if ind_id and hh_id:
                result[ind_id] = hh_id
        return result

    log.info(f"  HH member map: {len(ind_ids):,} individuals in {len(batches)} batches ...")
    with ThreadPoolExecutor(max_workers=_WORKERS) as ex:
        futures = {ex.submit(_one_batch, b): b for b in batches}
        for f in as_completed(futures):
            try:
                member_map.update(f.result())
            except Exception as e:
                _record_batch_failure("HH member", e)
    log.info(f"  HH member map: {len(member_map):,} memberships resolved")
    return member_map


def _map_household_ids_to_head_ids(cfg, hh_clref_ids):
    """
    Step 5 (Togo pattern): for each household ID find the head-of-household's individual ID.
    Queries household-member-index by Data.householdMember.householdClientReferenceId
    filtered to isHeadOfHousehold=True.
    Returns {hh_clref_id: head_individual_clref_id}.
    """
    if not hh_clref_ids:
        return {}
    batches  = [hh_clref_ids[i:i + _BATCH] for i in range(0, len(hh_clref_ids), _BATCH)]
    head_map = {}

    def _one_batch(batch):
        q = {
            "size": _BATCH,
            "query": {"bool": {"must": [
                {"terms": {"Data.householdMember.householdClientReferenceId.keyword": batch}},
                {"term":  {"Data.householdMember.isHeadOfHousehold": True}},
            ]}},
            "_source": [
                "Data.householdMember.householdClientReferenceId",
                "Data.householdMember.individualClientReferenceId",
            ],
        }
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_HH_MEMBER']}/_search",
            json=q, auth=cfg["es_auth"], verify=False, timeout=60,
        )
        r.raise_for_status()
        result = {}
        for h in r.json()["hits"]["hits"]:
            src    = h["_source"].get("Data", {}).get("householdMember", {})
            hh_id  = src.get("householdClientReferenceId", "")
            ind_id = src.get("individualClientReferenceId", "")
            if hh_id and ind_id:
                result[hh_id] = ind_id
        return result

    log.info(f"  HH head map: {len(hh_clref_ids):,} households in {len(batches)} batches ...")
    with ThreadPoolExecutor(max_workers=_WORKERS) as ex:
        futures = {ex.submit(_one_batch, b): b for b in batches}
        for f in as_completed(futures):
            try:
                head_map.update(f.result())
            except Exception as e:
                _record_batch_failure("HH head", e)
    log.info(f"  HH head map: {len(head_map):,} heads resolved")
    return head_map


def _read_target_sheet_url(url):
    """Read a target book held as a Google Sheet. Returns a DataFrame, or None
    if the URL is unparseable. Shared with analyze_itn so both drug families
    accept the same target_file forms."""
    import re

    import gspread
    from google.oauth2.service_account import Credentials

    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        log.error(f"Could not parse sheet ID from target book URL: "
                  f"{url} — TARGETS WILL BE ZERO and every coverage "
                  f"figure meaningless")
        return None
    sheet_id = m.group(1)
    gid_m = re.search(r"[#&]gid=(\d+)", url)

    from pipeline.config import _resolve_creds_path
    creds = Credentials.from_service_account_file(
        _resolve_creds_path(),
        scopes=["https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"])
    spreadsheet = gspread.Client(auth=creds).open_by_key(sheet_id)
    if gid_m:
        ws = next((w for w in spreadsheet.worksheets()
                   if str(w.id) == gid_m.group(1)), spreadsheet.sheet1)
    else:
        ws = spreadsheet.sheet1
    df = pd.DataFrame(ws.get_all_records())
    log.info(f"target book loaded from Google Sheet: {sheet_id} ({len(df)} rows)")
    return df


def _load_targets(cfg):
    from pipeline.core.drive import resolve_target_book
    csv_path = resolve_target_book(cfg)
    if not csv_path:
        log.error("no target book configured — all targets = 0, so "
                  "every coverage figure in this report will be 0% and meaningless — the report should not be shared until the target book is fixed")
        _degrade_run(TARGETS_ZERO + "no target book is configured. Set "
                     "target_file on the sheet row, or DST_TARGET_FOLDER_ID "
                     "for the deployment")
        return {}

    if csv_path.startswith("https://docs.google.com/spreadsheets/"):
        df = _read_target_sheet_url(csv_path)
        if df is None:
            _degrade_run(TARGETS_ZERO + "the target Google Sheet could not be "
                         "read: " + csv_path)
            return {}
    elif not os.path.exists(csv_path):
        log.error(f"target book not found: {csv_path} — all targets = 0, so "
                  f"every coverage figure in this report will be 0% and meaningless — the report should not be shared until the target book is fixed")
        _degrade_run(TARGETS_ZERO + "the target book was not found at "
                     + csv_path)
        return {}
    else:
        df = pd.read_csv(csv_path)
    if "facility_name" in df.columns and "individual_target" in df.columns:
        col_name, col_tgt = "facility_name", "individual_target"
    else:
        if len(df.columns) < 2:
            raise ValueError(
                f"target book has {len(df.columns)} column(s) {list(df.columns)} — "
                f"need facility_name and individual_target")
        col_name, col_tgt = df.columns[0], df.columns[1]
        # falling back to POSITION silently produced plausible-but-false coverage
        # for every facility when a book was re-exported with different headers
        log.error("target book lacks facility_name/individual_target; falling back "
                  "to positional columns %r (name) and %r (target) — verify these",
                  col_name, col_tgt)
    # Cumulative report measures against the FULL campaign target (undivided);
    # daily reports divide the total by campaign_days. divisor=1 keeps the overall target.
    divisor = 1 if cfg.get("cumulative") else cfg["campaign_days"]
    tmap = {}
    for _, row in df.iterrows():
        name = str(row[col_name]).strip().lower()
        try:
            itgt = float(row[col_tgt])
        except Exception:
            itgt = 0.0
        daily = round(itgt / divisor) if itgt > 0 else 0
        tmap[name] = {"4day": int(itgt), "daily": daily}
    log.info(f"  targets loaded: {len(tmap)} facilities")
    return tmap


def _coverage_band(treat_rate, daily_target, records, drug_type):
    # NOT REPORTED: zero records = the facility sent nothing at all (only the
    # target-book zero rows can have 0 — any ES-seen facility has >= 1 record)
    if records == 0:
        return "NOT REPORTED"
    # LOW ACTIVITY: <10 records regardless of coverage rate
    if records < 10:
        return "LOW ACTIVITY"
    if daily_target == 0:
        return "NO TARGET"
    # Coverage bands: HIGH >=95% | MODERATE 70-95% | LOW <70%
    if treat_rate >= 95:
        return "HIGH"
    if treat_rate >= 70:
        return "MODERATE"
    return "LOW"


def _aggregate_batch(task_hits, name_map, hh_name_map, fac_data, cfg):
    """
    Process one batch of task hits, updating the running fac_data dict in-place.
    Uses integer hash keys for dedup to keep memory ~7x lower than storing full tuples.
    """
    drug_type = cfg["drug_type"]
    min_age   = 3 if drug_type == "SPAQ" else 1

    for h in task_hits:
        doc  = h["_source"]["Data"]
        bh   = doc.get("boundaryHierarchy") or {}
        lga  = str(bh.get("lga",  "") or bh.get("district", "") or "").strip()
        ward = str(bh.get("ward", "") or bh.get("locality", "") or "").strip()
        fac  = str(bh.get("healthFacility", "") or bh.get("HEALTHFACILITY", "") or "").strip()
        if not fac:
            continue

        ind_id     = doc.get("individualId", "") or ""
        adm        = doc.get("administrationStatus", "") or ""
        add        = doc.get("additionalDetails") or {}
        # lat/lon: top-level Data.latitude (Nigeria states) with additionalDetails fallback (Togo)
        lat        = doc.get("latitude")  if doc.get("latitude")  not in (None, "") else add.get("latitude")
        lon        = doc.get("longitude") if doc.get("longitude") not in (None, "") else add.get("longitude")
        del_com    = str(add.get("deliveryComments") or doc.get("deliveryComments") or "").strip()
        # name maps are resolved per-task in run() and keyed by the task's clientReferenceId
        cref       = doc.get("clientReferenceId", "") or ""
        child_name = name_map.get(cref, "")
        hh_head    = hh_name_map.get(cref, "")

        age_raw = doc.get("age")
        try:
            age = int(float(age_raw)) if age_raw not in (None, "") else None
        except Exception:
            age = None

        qty_raw = doc.get("quantity")
        try:
            qty = int(float(qty_raw)) if qty_raw not in (None, "") else 0
        except Exception:
            qty = 0

        inelig_status = adm in ("BENEFICIARY_INELIGIBLE", "INELIGIBLE")
        ref_status    = adm == "BENEFICIARY_REFERRED"

        if fac not in fac_data:
            fac_data[fac] = dict(
                lga=lga, fac=fac, records=0, treated=0,
                drug1=0, drug2=0, absent=0, refused=0,
                ineligible=0, referred=0, died=0, migrated=0,
                redose=0, age_over59=0, age_zero=0,
                missing_hh=0, missing_child=0, missing_gender=0, duplicates=0,
                delivery_comments=0, missing_lat_lon=0,
                wards=set(), seen_keys=set(),
            )

        m = fac_data[fac]
        m["records"] += 1

        if ward:
            m["wards"].add(ward)
        if (lat is None or lat == "") or (lon is None or lon == ""):
            m["missing_lat_lon"] += 1
        if del_com:
            m["delivery_comments"] += 1
        if not child_name:
            m["missing_child"] += 1
        if not hh_head:
            m["missing_hh"] += 1
        # gender lives in Data.gender for most tenants but in additionalDetails
        # for the admin-console ones (Kebbi) - check BOTH. Reading only one side
        # reports 100% missing for every tenant that uses the other.
        if not (doc.get("gender") or add.get("gender")):
            m["missing_gender"] += 1

        if age is None:
            pass
        elif age == 0:
            m["age_zero"] += 1
        elif age > 59:
            m["age_over59"] += 1

        if   adm == "BENEFICIARY_ABSENT":                         m["absent"]    += 1
        elif adm == "BENEFICIARY_REFUSED":                        m["refused"]   += 1
        elif adm in ("BENEFICIARY_INELIGIBLE", "INELIGIBLE"):    m["ineligible"] += 1
        elif adm == "BENEFICIARY_REFERRED":                       m["referred"]  += 1
        elif adm == "BENEFICIARY_DIED":                           m["died"]      += 1
        elif adm == "BENEFICIARY_MIGRATED":                       m["migrated"]  += 1

        if adm == "VISITED" and qty >= 1:
            m["redose"] += 1

        is_treated = (
            age is not None and min_age <= age <= 59
            and not inelig_status and not ref_status
            and qty >= 1 and adm != "VISITED"
        )
        if is_treated:
            m["treated"] += 1
            if age is not None and age <= 11:
                m["drug2"] += 1
            else:
                m["drug1"] += 1

            # dedup TREATED records only: a redose (VISITED after SUCCESS) or a
            # status progression (ABSENT/REFUSED then treated) creates a second
            # record for the same child that is NOT a duplicate treatment. The
            # key carries no status/date, so checking every record counted those
            # legitimate sequences as duplicates and inflated the DQ metric.
            if child_name:
                dedup_key = f"{hh_head.lower()}|{child_name.lower()}|{ward.lower()}|{age}"
                if dedup_key in m["seen_keys"]:
                    m["duplicates"] += 1
                else:
                    m["seen_keys"].add(dedup_key)


def _build_facility_rows(fac_data, target_map, cfg):
    """Convert running fac_data dict into the final sorted results list.

    Rows are the UNION of facilities that reported today and the target book:
    a target-book facility with no records today becomes an explicit zero row
    (status NOT REPORTED = target present, no data synced; LGA shown as a
    dash — the CSV carries no LGA information). This keeps the report's
    Daily Target STANDARD — full campaign target / campaign_days —
    irrespective of which facilities synced. Without it the denominator
    silently shrank to the reporting subset (Chad Cycle 2: Day 1 target
    59,645 vs Day 2 83,340 from the same unchanged CSV, 118% "coverage").
    """
    matched = {f.lower() for f in fac_data}
    for name, tgt_entry in target_map.items():
        if name not in matched and tgt_entry.get("daily", 0) > 0:
            fac_data[name.upper()] = dict(
                lga="—", fac=name.upper(), records=0, treated=0,
                drug1=0, drug2=0, absent=0, refused=0,
                ineligible=0, referred=0, died=0, migrated=0,
                redose=0, age_over59=0, age_zero=0,
                missing_hh=0, missing_child=0, missing_gender=0, duplicates=0,
                delivery_comments=0, missing_lat_lon=0,
                wards=set(), seen_keys=set(),
            )

    results = []
    for fac, m in sorted(fac_data.items(), key=lambda x: (x[1]["lga"], x[0])):
        tgt_entry   = target_map.get(fac.lower(), {"4day": 0, "daily": 0})
        daily_tgt   = tgt_entry["daily"]
        records     = m["records"]
        treated     = m["treated"]
        not_treated = records - treated

        if daily_tgt > 0:
            rate = treated / daily_tgt * 100
            cov  = f"{rate:.1f}%"
        else:
            rate = 0.0
            cov  = "NO TARGET"

        flag = _coverage_band(rate, daily_tgt, records, cfg["drug_type"])

        results.append({
            "lga":               m["lga"],
            "fac":               fac,
            "daily_target":      daily_tgt,
            "records":           records,
            "treated":           treated,
            "not_treated":       not_treated,
            "drug1":             m["drug1"],
            "drug2":             m["drug2"],
            "coverage":          cov,
            "status":            flag,
            "absent":            m["absent"],
            "refused":           m["refused"],
            "ineligible":        m["ineligible"],
            "referred":          m["referred"],
            "died":              m["died"],
            "migrated":          m["migrated"],
            "redose":            m["redose"],
            "age_over59":        m["age_over59"],
            "age_zero":          m["age_zero"],
            "missing_hh":        m["missing_hh"],
            "missing_child":     m["missing_child"],
            "missing_gender":    m["missing_gender"],
            "duplicates":        m["duplicates"],
            "delivery_comments": m["delivery_comments"],
            "wards":             ", ".join(sorted(m["wards"])),
            "rate":              rate,
        })
    return results


def _fetch_secondary_counts(cfg, spec):
    """
    Count one secondary product's successful deliveries per facility.
    Filters: productName = spec['name'], ADMINISTRATION_SUCCESS, age in
    [spec['age_min'], spec['age_max']], plus the campaign scope filters.
    Called once per entry in cfg['secondary_products'] (Kebbi ORS-Zinc = 1 entry,
    Bauchi VAS = 2 entries: Red VAS 12-59, Blue VAS 6-11).
    """
    product    = spec["name"]
    amin, amax = spec["age_min"], spec["age_max"]
    date_field = cfg.get("task_date_field", "taskDates")
    filters = [
        {"range": {f"Data.{date_field}": {"gte": cfg["GTE"], "lte": cfg["LTE"]}}},
        {"term":  {"Data.productName.keyword": product}},
        {"term":  {"Data.administrationStatus.keyword": "ADMINISTRATION_SUCCESS"}},
    ] + _build_campaign_filters(cfg)
    # Age filter only when a band is configured; VAS is counted by productName alone.
    if amin is not None and amax is not None:
        filters.append({"range": {"Data.age": {"gte": amin, "lte": amax}}})
    query = {
        "size": _BATCH,
        "query": {"bool": {"filter": filters}},
        "_source": ["Data.boundaryHierarchy", "Data.age"],
    }
    fac_counts = defaultdict(lambda: {"lga": "", "count": 0})
    for batch in scroll_batches(cfg["es_url"], cfg["ES_INDEX_TASK"], query,
                                 cfg["es_auth"], f"secondary-{product}"):
        for h in batch:
            doc = h["_source"]["Data"]
            bh  = doc.get("boundaryHierarchy") or {}
            fac = str(bh.get("healthFacility", "") or bh.get("HEALTHFACILITY", "") or "").strip()
            lga = str(bh.get("lga",            "") or bh.get("district",       "") or "").strip()
            if fac:
                fac_counts[fac]["lga"]    = lga
                fac_counts[fac]["count"] += 1

    return sorted(
        [{"lga": v["lga"], "fac": k, "count": v["count"]}
         for k, v in fac_counts.items()],
        key=lambda x: (x["lga"], x["fac"]),
    )


def _write_secondary_tab(wb, rows, cfg, spec):
    """Write one secondary-product tab: LGA | Health Facility | <label> Count."""
    label      = spec["label"]
    amin, amax = spec["age_min"], spec["age_max"]
    product    = label
    ws      = wb.create_sheet(spec["tab"])
    total   = sum(r["count"] for r in rows)
    band_txt = f"Age {amin}-{amax} months" if amin is not None and amax is not None else "All ages"
    banner  = f"{label} Distribution — Day {cfg['DAY']}  |  Total: {total:,}  |  {band_txt}"

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = banner
    c.fill  = BANNER_FILL
    c.font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for ci, h in enumerate(["LGA", "Health Facility", f"{product} Count"], 1):
        cell = ws.cell(row=2, column=ci, value=h)
        style_cell(cell, fill=HDR_FILL, bold=True, color="FFFFFF")

    for ri, r in enumerate(rows, 1):
        for ci, val in enumerate([r["lga"], r["fac"], r["count"]], 1):
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            style_cell(cell, fill=WHITE_FILL)

    # totals row
    tr = len(rows) + 3
    for ci, val in enumerate(["", "TOTAL", total], 1):
        cell = ws.cell(row=tr, column=ci, value=val)
        style_cell(cell, fill=TOTAL_FILL, bold=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18


# ── Excel writing ──────────────────────────────────────────────────────────────

def _column_headers(drug_type, cumulative=False):
    d1 = "SPAQ2 (12-59m)" if drug_type == "SPAQ" else "AZM 12-59m"
    d2 = "SPAQ1 (3-11m)"  if drug_type == "SPAQ" else "AZM 1-11m"
    tgt = "Campaign Target" if cumulative else "Daily Target"
    return [
        "#", "LGA", "Health Facility", tgt, "Records", "Treated",
        "Not Treated", d1, d2, "Coverage %", "Status",
        "Absent", "Refused", "Ineligible", "Referred", "Died", "Migrated",
        "Redose", "Age>59", "Age=0", "Missing HH", "Missing Child", "Missing Gender",
        "Duplicates", "Delivery Comments", "Wards",
    ]


def _row_values(r, idx):
    return [
        idx, r["lga"], r["fac"], r["daily_target"], r["records"], r["treated"],
        r["not_treated"], r["drug1"], r["drug2"], r["coverage"], r["status"],
        r["absent"], r["refused"], r["ineligible"], r["referred"], r["died"],
        r["migrated"], r["redose"], r["age_over59"], r["age_zero"],
        r["missing_hh"], r["missing_child"], r["missing_gender"], r["duplicates"],
        r["delivery_comments"], r["wards"],
    ]


def _totals_row(rows, drug_type):
    def s(k): return sum(r[k] for r in rows)
    total_tgt  = s("daily_target")
    total_rec  = s("records")
    total_tr   = s("treated")
    cov = f"{total_tr/total_tgt*100:.1f}%" if total_tgt else "N/A"
    d1 = "SPAQ2 (12-59m)" if drug_type == "SPAQ" else "AZM 12-59m"
    d2 = "SPAQ1 (3-11m)"  if drug_type == "SPAQ" else "AZM 1-11m"
    return {
        "lga": "", "fac": "GRAND TOTAL",
        "daily_target": total_tgt, "records": total_rec, "treated": total_tr,
        "not_treated": s("not_treated"), "drug1": s("drug1"), "drug2": s("drug2"),
        "coverage": cov, "status": "",
        "absent": s("absent"), "refused": s("refused"), "ineligible": s("ineligible"),
        "referred": s("referred"), "died": s("died"), "migrated": s("migrated"),
        "redose": s("redose"), "age_over59": s("age_over59"), "age_zero": s("age_zero"),
        "missing_hh": s("missing_hh"), "missing_child": s("missing_child"),
        "missing_gender": s("missing_gender"),
        "duplicates": s("duplicates"), "delivery_comments": s("delivery_comments"),
        "wards": "", "rate": 0,
    }


def _write_facility_tab(ws, rows, headers, drug_type, cfg, banner_text):
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    # row 1: banner
    ws.merge_cells(f"A1:{last_col}1")
    banner = ws["A1"]
    banner.value = banner_text
    banner.fill  = BANNER_FILL
    banner.font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # row 2: headers
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        style_cell(cell, fill=HDR_FILL, bold=True, align="center", color="FFFFFF")

    # data rows
    for ri, r in enumerate(rows, 1):
        vals = _row_values(r, ri)
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            style_cell(cell, fill=WHITE_FILL, align="center")
            # color Status and Coverage % cols
            if ci == 10:  # Coverage %
                pass
            if ci == 11:  # Status
                flag_col = FLAG_COLOR.get(str(val), "000000")
                cell.font = Font(bold=True, color=flag_col, size=9, name="Calibri")

    # totals row
    if rows:
        tot = _totals_row(rows, drug_type)
        tot_row = len(rows) + 3
        tot_vals = _row_values(tot, "")
        for ci, val in enumerate(tot_vals, 1):
            cell = ws.cell(row=tot_row, column=ci, value=val)
            style_cell(cell, fill=TOTAL_FILL, bold=True, align="center")

    # freeze, filter, col widths
    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{last_col}2"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    for ci in range(4, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.column_dimensions[get_column_letter(ncols)].width = 30  # Wards


# ── public entry point ─────────────────────────────────────────────────────────

BANDS = ["LOW", "MODERATE", "HIGH", "NO TARGET", "LOW ACTIVITY", "NOT REPORTED"]


def _build_task_queries(cfg):
    """One scroll query per administrationStatus family, sharing date + campaign filters."""
    source = [
        "Data.boundaryHierarchy", "Data.age", "Data.gender", "Data.individualId",
        "Data.clientReferenceId", "Data.projectBeneficiaryClientReferenceId",
        "Data.quantity", "Data.administrationStatus",
        "Data.latitude", "Data.longitude", "Data.additionalDetails",
    ]
    date_field = cfg.get("task_date_field", "taskDates")
    base = [{"range": {f"Data.{date_field}": {"gte": cfg["GTE"], "lte": cfg["LTE"]}}}]
    base += _build_campaign_filters(cfg)

    treatment = base + [
        {"terms": {"Data.administrationStatus.keyword": ["ADMINISTRATION_SUCCESS", "VISITED"]}},
    ]
    if cfg.get("dose_index_filter"):
        treatment.append({"term": {"Data.additionalDetails.doseIndex.keyword": "1"}})

    nonadmin = base + [
        {"terms": {"Data.administrationStatus.keyword": [
            "BENEFICIARY_INELIGIBLE", "INELIGIBLE", "BENEFICIARY_REFERRED",
            "BENEFICIARY_DIED", "BENEFICIARY_ABSENT", "BENEFICIARY_MIGRATED",
            "BENEFICIARY_REFUSED",
        ]}},
    ]
    return [
        ("task-treatment", {"size": _BATCH, "query": {"bool": {"filter": treatment}}, "_source": source}),
        ("task-nonadmin",  {"size": _BATCH, "query": {"bool": {"filter": nonadmin}},  "_source": source}),
    ]


def _resolve_batch_names(cfg, docs):
    """Resolve child and household-head names for one batch of task docs.

    Names come only from the individual index, via the chain:
    project-task -> project-beneficiary -> household-member -> individual.
    Returns (child_name_map, head_name_map) keyed by task clientReferenceId.
    """
    pb_refs = list({d.get("projectBeneficiaryClientReferenceId") or "" for d in docs
                    if d.get("projectBeneficiaryClientReferenceId")})
    pb2benef = _map_beneficiary_refs_to_individual_ids(cfg, pb_refs)

    benef_ids = list({v for v in pb2benef.values() if v})
    benef2hh = _map_individual_ids_to_household_ids(cfg, benef_ids)

    hh_ids = list({v for v in benef2hh.values() if v})
    hh2head = _map_household_ids_to_head_ids(cfg, hh_ids)

    ind_names = _fetch_individual_names(
        cfg, list(set(benef_ids) | {v for v in hh2head.values() if v}))

    name_map, hh_name_map = {}, {}
    for d in docs:
        cr = d.get("clientReferenceId") or ""
        if not cr:
            continue
        benef = pb2benef.get(d.get("projectBeneficiaryClientReferenceId") or "", "")
        head = hh2head.get(benef2hh.get(benef, ""), "")
        name_map[cr] = ind_names.get(benef, "")
        hh_name_map[cr] = ind_names.get(head, "")
    return name_map, hh_name_map


def collect(cfg):
    """Stage 1 (all ES I/O): stream task docs, resolve names, aggregate per facility.

    Returns (facility_rows, secondary_rows_by_label, records_processed).
    """
    target_map = _load_targets(cfg)
    fac_data = {}
    total_processed = 0
    for label, query in _build_task_queries(cfg):
        for batch in scroll_batches(cfg["es_url"], cfg["ES_INDEX_TASK"], query, cfg["es_auth"], label):
            docs = [h["_source"]["Data"] for h in batch]
            name_map, hh_name_map = _resolve_batch_names(cfg, docs)
            _aggregate_batch(batch, name_map, hh_name_map, fac_data, cfg)
            total_processed += len(batch)

    rows = _build_facility_rows(fac_data, target_map, cfg)
    log.info(f"[analyze:collect] {total_processed:,} records -> {len(rows)} facilities")

    secondary = {}
    for spec in cfg.get("secondary_products", []):
        sec_rows = _fetch_secondary_counts(cfg, spec)
        secondary[spec["label"]] = sec_rows
        log.info(f"[analyze:collect] {spec['label']}: {sum(r['count'] for r in sec_rows):,} "
                 f"records across {len(sec_rows)} facilities")
    return rows, secondary, total_processed


def render(cfg, rows, secondary):
    """Stage 2 (pure, no I/O beyond the output file): build the performance workbook."""
    drug_type = cfg["drug_type"]
    headers = _column_headers(drug_type, cumulative=cfg.get("cumulative", False))

    total_target = sum(r["daily_target"] for r in rows)
    if cfg.get("cumulative"):
        banner_text = (
            f"Overall Campaign Target: {total_target:,}  |  "
            f"Cumulative Days 1-{cfg['DAY']}  ({cfg['START_LABEL']} to {cfg['END_LABEL']})"
        )
    else:
        banner_text = (
            f"Campaign Target: {total_target * cfg['campaign_days']:,}  |  "
            f"Daily Target (Day {cfg['DAY']}): {total_target:,}  "
            f"(= Campaign Target ÷ {cfg['campaign_days']} campaign days; "
            f"all targets/coverage in this sheet are vs the DAILY target)"
        )

    wb = Workbook()
    wb.remove(wb.active)
    _write_facility_tab(wb.create_sheet("ALL FACILITIES"), rows, headers, drug_type, cfg, banner_text)
    for band in BANDS:
        band_rows = [r for r in rows if r["status"] == band]
        _write_facility_tab(wb.create_sheet(band), band_rows, headers, drug_type, cfg, banner_text)
    for spec in cfg.get("secondary_products", []):
        _write_secondary_tab(wb, secondary.get(spec["label"], []), cfg, spec)

    out = cfg["perf_xlsx"]
    wb.save(out)
    log.info(f"[analyze:render] saved -> {out}  ({len(rows)} facilities)")
    return out


def run(cfg):
    log.info(f"[analyze] {cfg['state_name']} Day {cfg['DAY']} — streaming task docs ...")
    rows, secondary, processed = collect(cfg)
    if not processed:
        # A green report over zero documents is the most dangerous output this
        # pipeline can produce: indistinguishable from a real quiet day. It has
        # happened - a repo .env overriding ES_INDEX_PREFIX put every query on
        # an index that exists but holds another tenant's data.
        _degrade_run(
            f"ZERO task documents matched. Every figure in this report is 0 and "
            f"is NOT a measurement. Check, in this order: the index actually "
            f"queried (tenant={cfg.get('tenant')}, ES_INDEX_PREFIX="
            f"{os.getenv('ES_INDEX_PREFIX', '<unset: tenant-prefixed>')!r}); the "
            f"date window {cfg.get('GTE')} to {cfg.get('LTE')} on field "
            f"{cfg.get('task_date_field', 'taskDates')}; and campaign_number / "
            f"cycle_index if task_campaign_filter is TRUE")
    save_checkpoint(cfg, "analyze", {
        "processed": processed, "rows": rows, "secondary": secondary,
    })
    return render(cfg, rows, secondary)


def rerun_from_checkpoint(cfg):
    """Rebuild the performance Excel from the saved checkpoint — no ES access needed."""
    state = load_checkpoint(cfg, "analyze")
    return render(cfg, state["rows"], state["secondary"])
