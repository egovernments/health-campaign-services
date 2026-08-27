"""
analyze_itn.py — ITN/LLIN (bed-net) household-based aggregation → performance Excel

Separate module from analyze.py by design: SPAQ/AZM's core computation (is_treated
requires a non-null Data.age; one child = one dose) does not apply to ITN campaigns
(beneficiaryType=HOUSEHOLD, Data.age is always null, one household = N nets based on
household size). Shares only the generic primitives in pipeline.core (ES scroll,
Excel styling) plus analyze.py's name-resolution lookups — analyze.py's aggregation
logic is never imported and never modified by this file.

Confirmed on live chad ES (tenant=chad, campaign CMP-2026-06-03-000312, LLIN_phase2_
tchad2026-final-3):
  - Data.additionalDetails.projectReferenceId (NOT Data.campaignNumber) scopes the
    campaign — this tenant's task docs have no top-level campaignNumber field.
  - Facility tier field is Data.boundaryHierarchy.sppSfd / boundaryHierarchyCode.sppSfd
    (NOT healthFacility/HEALTHFACILITY like SPAQ/AZM tenants).
  - Data.householdId, Data.quantity, Data.memberCount all aggregate correctly
    (cardinality / sum) — validated against live ES.

AGGREGATION GRAIN: LGA, not health facility. Reasons:
  1. There are 1,000+ health facilities for this campaign, but only ~40 LGAs — a
     facility-level target was never actually available, and matching 1,000+
     facility names 1:1 is far more fragile than matching ~40 LGA codes.
  2. Real target data for this tenant only exists at LGA grain (confirmed via a
     live DB query — see below).
Facility-level detail (records, DQ, low-activity) is STILL scrolled and kept — it
just doesn't carry a target/coverage/status anymore, since no facility-level
target exists. LGA rows carry the real target/coverage/status.

TARGET SOURCE, confirmed via live DB + ES cross-check:
  - chad.project_target rows are keyed to chad.project_address.boundary, which for
    boundarytype='LGA' is a CODE like "ADMIN_TC_16_03_ADRE" — NOT the plain LGA
    name ("ADRE") that ES's Data.boundaryHierarchy.district field carries.
  - Confirmed via live ES query: every task doc with boundaryHierarchy.district =
    "ADRE" carries boundaryHierarchyCode.district = "ADMIN_TC_16_03_ADRE" — an
    exact match to the DB's chad.project_target boundary code for that LGA. So
    matching MUST be done on the CODE (boundaryHierarchyCode.district), never on
    the plain name — matching on name would silently produce zero matches.
  - DB query used (chad tenant), beneficiarytype IN ('INDIVIDUAL', 'HOUSEHOLD'):
        SELECT pa.boundary AS lga_code,
               SUM(CASE WHEN pt.beneficiarytype='INDIVIDUAL' THEN pt.targetno ELSE 0 END) AS population_target,
               SUM(CASE WHEN pt.beneficiarytype='HOUSEHOLD'  THEN pt.targetno ELSE 0 END) AS household_target
        FROM chad.project p
        JOIN chad.project_target pt  ON pt.projectid = p.id AND p.isdeleted = false
        JOIN chad.project_address pa ON pa.projectid = p.id AND pa.boundarytype = 'LGA'
        WHERE p.referenceid = '<campaign reference id>'
          AND pt.beneficiarytype IN ('INDIVIDUAL', 'HOUSEHOLD')
        GROUP BY pa.boundary;
  - Net/ITN target (beneficiarytype='PRODUCT') confirmed present too — same query
    with beneficiarytype IN (..., 'PRODUCT') returns real net_target values that
    exactly match an independent cross-check against the campaign's microplan
    Excel (ADRE: 202,850 both ways). Target CSV itself is NOT sourced from the
    microplan (explicitly ruled out) — that was only used once as a cross-check.

DATE SCOPING — matches analyze.py's mechanism exactly:
  - Confirmed via live ES: Data.taskDates exists on chad's task docs (same field
    name AND format SPAQ already defaults to, e.g. "2026-07-04"), and a real
    GTE/LTE range query against it returns a correct, non-zero, verified count
    (22 records for 2026-07-04) — i.e. it behaves as a proper ES date field, not
    just a keyword string, so the same range-filter pattern analyze.py already
    uses works here without any change.
  - Every run is now scoped to cfg["GTE"]/cfg["LTE"] (config.py already computes
    these for every row, generically — bounds exactly "today" for a daily run, or
    the whole campaign for a --cumulative run), via the SAME field name default
    ("taskDates") and the SAME cfg.get("task_date_field", "taskDates") override
    convention as analyze.py. This means cfg["perf_xlsx"]/itn_history/ snapshots
    are now genuinely TODAY-ONLY (or campaign-wide, in cumulative mode) — no
    longer an always-cumulative-to-date total. report_itn.py sums multiple days'
    snapshots together for the cumulative view, exactly like report.py's
    _load_all_days_perf, instead of subtracting cumulative snapshots.
  - Target division: this campaign's target CSV holds the FULL campaign-length
    target (not a pre-divided daily allocation). For a daily (non-cumulative) run,
    _load_targets_itn's values are divided by the real campaign length in days
    (cfg["campaign_start"] to cfg["campaign_end"], NOT cfg["campaign_days"] —
    config.py defaults that to 4 when a sheet row doesn't set it) before being
    used as each LGA's Status/Coverage denominator — same "Total Campaign Target
    = Daily Target × Campaign Days" relationship analyze.py's own daily mode
    uses. Cumulative runs use the full undivided target directly.
"""
import logging
import os
from datetime import datetime, timezone

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from pipeline.analyze import (_map_household_ids_to_head_ids,
                             _fetch_individual_names,
                             _degrade_run, TARGETS_ZERO)
from pipeline.core.es import scroll_batches
from pipeline.core.excel import (
    BANNER_FILL, FLAG_COLOR, HDR_FILL, TOTAL_FILL, WHITE_FILL, style_cell,
)

log = logging.getLogger(__name__)

_BATCH = 5000

# Duplicate-distribution matrix bucket keys (opt-in — see
# config._dup_matrix_default: dup_matrix sheet cell, then DST_DUP_MATRIX).
# su/du = same/different user; sd/dd = same/different day, always relative to
# the household's ORIGINAL (earliest) record. Duplicates with a missing user,
# date or head name are excluded, never guessed (run() warns when the excluded
# share is large). None on every row means "not measured this run" and renders
# as an empty cell — deliberately distinct from a real 0.
_DUP_KEYS = ("dup_su_sd", "dup_su_dd", "dup_du_sd", "dup_du_dd")


# ── ES helpers ─────────────────────────────────────────────────────────────────

def _campaign_filter(cfg):
    """
    ITN campaign scoping. Confirmed for tenant 'chad': the task index carries the
    campaign reference at Data.additionalDetails.projectReferenceId, NOT at a
    top-level Data.campaignNumber field (unlike the SPAQ/AZM admin-console tenants).
    """
    if not cfg.get("campaign_number"):
        raise ValueError("campaign_number is required for ITN reporting")
    return {"term": {"Data.additionalDetails.projectReferenceId.keyword": cfg["campaign_number"]}}


def _date_filter(cfg):
    """
    Same mechanism as analyze.py: a GTE/LTE range filter on Data.taskDates
    (override via cfg["task_date_field"], same convention), bounding this run to
    cfg["GTE"]/["LTE"] — which config.py already computes generically for every
    row (today's date for a daily run, the whole campaign span for --cumulative).
    Confirmed via live ES: Data.taskDates on chad's task docs is a
    real ES date field (range query returns a correct non-zero count), not just
    a keyword string, so this range filter works the same way it does for SPAQ.
    """
    date_field = cfg.get("task_date_field", "taskDates")
    return {"range": {f"Data.{date_field}": {"gte": cfg["GTE"], "lte": cfg["LTE"]}}}


def _fetch_facility_rows(cfg):
    """
    Scroll-based (NOT a pure ES aggregation) — required for two reasons proven this
    session:
    1. Resolving the household head's real name needs the SAME 2-hop join
       analyze.py uses for SPAQ/AZM (household-member-index -> individual-index).
       A single field-exists check on Data.additionalDetails.familyNameOfIndividual
       was tried and rejected — that field is a device-entered value, sometimes
       incomplete (confirmed: "Dorsso" vs the individual-index's real
       "Hadjidjad Dorsso" for the same person), and the project's own Name
       Resolution Rule says names must never be trusted from additionalDetails.
    2. Nets/population must be counted ONCE per distinct household, not summed
       across every task record. Sync-retry duplicates are real (confirmed: one
       facility had 197 records but only 62 distinct households) — summing
       Data.quantity/Data.memberCount across all of them would triple-count.
       Duplicate volume is still tracked, just as its own DQ metric (dup_records),
       not folded into the coverage numbers.

    Returns (facility_rows, dup_events): facility-level rows (still the natural ES
    aggregation unit — a task doc belongs to one facility), each carrying its LGA
    name + LGA code so the caller can roll them up to LGA grain for
    target-matching; and, ONLY when cfg["dup_matrix"] is on, one event dict per
    task record (date/ts/user/fac/householdId/identity) for the duplicate-matrix
    classifier ([] otherwise — zero extra memory for every existing deployment).
    """
    fac_data       = {}   # facility_code -> accumulator dict
    seen_hh_by_fac = {}   # facility_code -> set of already-counted householdIds
    hh_head_ind_map = {}  # householdId -> head's individual clientReferenceId (cached)
    head_name_map   = {}  # individual clientReferenceId -> resolved name (cached)

    collect_events = bool(cfg.get("dup_matrix"))
    dup_events     = []   # one dict per record — see _classify_duplicates
    event_errors   = 0
    date_field     = cfg.get("task_date_field", "taskDates")

    _source = [
        "Data.boundaryHierarchy", "Data.boundaryHierarchyCode", "Data.householdId",
        "Data.quantity", "Data.memberCount", "Data.latitude", "Data.longitude",
        "Data.deliveryComments", "Data.additionalDetails.manualCodes",
        "Data.additionalDetails.codesScanned",
    ]
    if collect_events:
        # Acting-user, record-date and household-identity fields, fetched only
        # when the matrix is on. See _user_and_ts() for the createdBy path
        # priority (chad carries it at top-level Data.createdBy) and
        # _identity() for why the household is matched on name+members+village
        # rather than householdId.
        _source += [
            f"Data.{date_field}",
            "Data.createdBy", "Data.createdTime",
            "Data.clientAuditDetails.createdBy", "Data.clientAuditDetails.createdTime",
            "Data.auditDetails.createdBy",
            "Data.additionalDetails.name", "Data.additionalDetails.familyNameOfIndividual",
            "Data.additionalDetails.memberCount", "Data.additionalDetails.administrativeArea",
        ]
    query = {
        "size": _BATCH,
        "query": {
            "bool": {
                "filter": [
                    _campaign_filter(cfg),
                    _date_filter(cfg),
                    {"term": {"Data.administrationStatus.keyword": "ADMINISTRATION_SUCCESS"}},
                ]
            }
        },
        "_source": _source,
    }

    total_processed = 0
    for batch in scroll_batches(cfg["es_url"], cfg["ES_INDEX_TASK"], query, cfg["es_auth"], "itn-task"):
        docs = [h["_source"]["Data"] for h in batch]

        # Resolve any new households' head-individual mapping seen in this batch.
        batch_hh_ids = list({d.get("householdId") for d in docs if d.get("householdId")})
        new_hh_ids   = [h for h in batch_hh_ids if h not in hh_head_ind_map]
        if new_hh_ids:
            hh_head_ind_map.update(_map_household_ids_to_head_ids(cfg, new_hh_ids))

        # Resolve any new heads' names seen in this batch.
        batch_head_ids = list({hh_head_ind_map[h] for h in batch_hh_ids if h in hh_head_ind_map})
        new_head_ids   = [i for i in batch_head_ids if i not in head_name_map]
        if new_head_ids:
            head_name_map.update(_fetch_individual_names(cfg, new_head_ids))

        for doc in docs:
            bh  = doc.get("boundaryHierarchy") or {}
            bhc = doc.get("boundaryHierarchyCode") or {}
            province = str(bh.get("province", "") or "").strip()
            lga      = str(bh.get("district", "") or "").strip()
            lga_code = str(bhc.get("district", "") or "").strip()
            fac_name = str(bh.get("sppSfd", "") or "").strip()
            fac_code = str(bhc.get("sppSfd", "") or "").strip()
            if not fac_code:
                continue

            hh_id   = doc.get("householdId") or ""
            lat     = doc.get("latitude")
            lon     = doc.get("longitude")
            add     = doc.get("additionalDetails") or {}
            if collect_events and hh_id:
                # Guarded per record: this runs on the MAIN scroll path, so a
                # malformed doc must cost one matrix event (undercount, warned
                # below), never the legacy report.
                try:
                    user, created_ts = _user_and_ts(doc)
                    given   = str(add.get("name") or "").strip()
                    family  = str(add.get("familyNameOfIndividual") or "").strip()
                    members = doc.get("memberCount")
                    if members in (None, ""):
                        members = add.get("memberCount")
                    try:
                        members = int(float(members))
                    except (ValueError, TypeError):
                        members = None
                    area = str(add.get("administrativeArea") or "").strip()
                    dup_events.append({
                        "date": _norm_date(doc.get(date_field)), "ts": created_ts,
                        "user": user, "fac": fac_code, "hh": hh_id,
                        "ident": _identity(f"{given} {family}", members, area or fac_code),
                        "head": f"{given} {family}".strip(), "family": family,
                        "members": members, "area": area,
                    })
                except Exception:
                    event_errors += 1
            try:
                manual = int(float(add.get("manualCodes") or 0))
            except (ValueError, TypeError):
                manual = 0
            try:
                scanned = int(float(add.get("codesScanned") or 0))
            except (ValueError, TypeError):
                scanned = 0

            if fac_code not in fac_data:
                fac_data[fac_code] = dict(
                    province=province, lga=lga, lga_code=lga_code,
                    facility_name=fac_name or fac_code,
                    records=0, households_visited=0, nets_distributed=0, population_covered=0,
                    missing_hh_head=0, missing_gps=0,
                    manual_codes=0, scanned_codes=0, missing_codes=0,
                )
                seen_hh_by_fac[fac_code] = set()

            m = fac_data[fac_code]
            m["records"] += 1
            if (lat is None or lat == "") or (lon is None or lon == ""):
                m["missing_gps"] += 1
            m["manual_codes"]  += manual
            m["scanned_codes"] += scanned
            # DQ: a net was distributed but NEITHER scanned NOR manually recorded a
            # code for it — the net has zero barcode documentation, untraceable in
            # inventory. Distinct from the manual-vs-scanned RATIO already tracked.
            if manual == 0 and scanned == 0:
                m["missing_codes"] += 1

            if hh_id and hh_id not in seen_hh_by_fac[fac_code]:
                # First time seeing this household at this facility — count its
                # nets/population/head-name ONCE, deduped against sync-retry repeats.
                seen_hh_by_fac[fac_code].add(hh_id)
                m["households_visited"] += 1
                try:
                    m["nets_distributed"] += int(float(doc.get("quantity") or 0))
                except (ValueError, TypeError):
                    pass
                try:
                    m["population_covered"] += int(float(doc.get("memberCount") or 0))
                except (ValueError, TypeError):
                    pass
                head_ind  = hh_head_ind_map.get(hh_id, "")
                head_name = head_name_map.get(head_ind, "") if head_ind else ""
                if not head_name:
                    m["missing_hh_head"] += 1

        total_processed += len(batch)

    if event_errors:
        log.warning(f"[analyze_itn] dup matrix: {event_errors:,} records skipped "
                    f"(malformed fields) — matrix will undercount by that many")
    log.info(f"[analyze_itn] {total_processed:,} records processed across {len(fac_data)} facilities")
    if not total_processed:
        # Same hazard as the SPAQ path: a complete, green ITN report over no data.
        _degrade_run(
            f"ZERO task documents matched. Every net and household figure in "
            f"this ITN report is 0 and is NOT a measurement. Check the index "
            f"queried (tenant={cfg.get('tenant')}, ES_INDEX_PREFIX="
            f"{os.getenv('ES_INDEX_PREFIX', '<unset: tenant-prefixed>')!r}), the "
            f"window {cfg.get('GTE')} to {cfg.get('LTE')}, and project_type_id / "
            f"campaign_number for this campaign")
    return [{"facility_code": code, **m} for code, m in fac_data.items()], dup_events


# ── duplicate-distribution matrix (opt-in: sheet column dup_matrix=TRUE) ──────

def _norm_date(v):
    """
    Normalise a task-date value to 'YYYY-MM-DD'. Stored values are date
    strings, but ES can also hand back epoch-ms and mapping quirks can wrap
    values in a list — all normalised; missing/empty -> "".
    """
    if isinstance(v, (list, tuple)):
        v = v[0] if v else None
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float)):
        return datetime.fromtimestamp(v / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
    return str(v)[:10]


def _identity(raw_name, members, scope):
    """
    Physical-household identity key for duplicate matching: normalised head
    name + member count, scoped to the village (additionalDetails.
    administrativeArea; facility code as fallback when blank). This is the
    only duplicate key — householdId is display-only, because a re-served
    household is usually re-registered under a brand-new householdId, which
    an ID-only match can never connect.
      - Name comes from the task doc's additionalDetails (name +
        familyNameOfIndividual), used purely as a match key — it is the only
        name present on both today's and history docs without a per-record
        2-hop join. Display names elsewhere still follow the Name Resolution
        Rule.
      - Village scoping is load-bearing: common head names collide heavily at
        any wider scope (chad has ~15,000 distinct households headed by
        "mahamat" alone; an LGA-scoped key merged dozens of real families per
        name).
      - Limits: a repeat spelled differently or recorded under a different
        village name will not match (undercount); two same-name same-size
        families in one village can still match (overcount, rare) — the DUP
        sheets carry name + village so a human can adjudicate.
    Returns None when no name is present — such records cannot be matched.
    """
    norm = " ".join(str(raw_name or "").lower().split())
    if not norm:
        return None
    m = "?" if members is None else str(members)
    return f"{norm}|{m}|{' '.join(str(scope or '').lower().split())}"


def _user_and_ts(d):
    """
    Acting-user UUID + creation timestamp from a task doc's Data payload.
    Path priority: clientAuditDetails.createdBy (device actor, tenants that
    carry it) -> top-level Data.createdBy (chad's shape, verified on live ES)
    -> auditDetails.createdBy. Returns ("", 0) when no path is populated —
    such records are excluded from the matrix, never guessed.
    Defensive against non-dict audit shapes: a malformed doc must degrade to
    ("", 0), not raise (this runs on the main scroll path).
    """
    caud = d.get("clientAuditDetails")
    aud  = d.get("auditDetails")
    caud = caud if isinstance(caud, dict) else {}
    aud  = aud if isinstance(aud, dict) else {}
    user = str(caud.get("createdBy") or d.get("createdBy") or aud.get("createdBy") or "").strip()
    try:
        ts = int(caud.get("createdTime") or d.get("createdTime") or 0)
    except (ValueError, TypeError):
        ts = 0
    return user, ts


def _fetch_dup_history(cfg, today_events):
    """
    History side of the duplicate matrix: fetch earlier ADMINISTRATION_SUCCESS
    records (campaign start up to, but excluding, this window's GTE) from the
    same task index for households seen in this run's window — the daily
    window never widens; cross-day evidence is a batched-terms lookup,
    scrolled per batch.

    Probes by familyNameOfIndividual (values seen today), which catches both
    re-syncs and re-registrations under a brand-new householdId. The terms
    prefilter is exact-match only; the real decision happens client-side via
    _identity, so a name hit with a different member count or village never
    groups.
    Limits: a repeat whose name was entered differently on the earlier day,
    or whose familyNameOfIndividual is blank (given name only), cannot be
    probed and its cross-day repeats are missed — undercount, never a false
    match. Same-day repeats (both inside the window) still classify.
    Returns a list of event dicts (fac=None marks them as history — never
    themselves counted as duplicates).
    """
    if not today_events:
        return []
    date_field = cfg.get("task_date_field", "taskDates")
    # The same name is stored in several raw shapes ("mahamat", "Mahamat",
    # "mahamat ") and the terms prefilter is exact — probe the common case/
    # trailing-space variants; client-side _identity decides the real match.
    fam_raw = {e["family"] for e in today_events if e.get("family")}
    variants = set()
    for f in fam_raw:
        for v in (f, f.lower(), f.upper(), f.capitalize(), f.title()):
            variants.add(v)
            variants.add(v + " ")
    families = sorted(variants)

    base_filters = [
        _campaign_filter(cfg),
        {"term": {"Data.administrationStatus.keyword": "ADMINISTRATION_SUCCESS"}},
        {"range": {f"Data.{date_field}": {
            "gte": str(cfg["campaign_start"])[:10], "lt": cfg["GTE"]}}},
    ]
    _source = [
        "Data.householdId", f"Data.{date_field}",
        "Data.createdBy", "Data.createdTime",
        "Data.clientAuditDetails.createdBy", "Data.clientAuditDetails.createdTime",
        "Data.auditDetails.createdBy",
        "Data.additionalDetails.name", "Data.additionalDetails.familyNameOfIndividual",
        "Data.additionalDetails.memberCount", "Data.memberCount",
        "Data.additionalDetails.administrativeArea", "Data.boundaryHierarchyCode.sppSfd",
    ]

    history, seen_ids = [], set()

    def _run_lookup(field, values, label):
        batches = [values[i:i + _BATCH] for i in range(0, len(values), _BATCH)]
        for bi, batch in enumerate(batches, 1):
            query = {
                "size": _BATCH,
                "query": {"bool": {"filter": base_filters + [{"terms": {field: batch}}]}},
                "_source": _source,
            }
            for page in scroll_batches(cfg["es_url"], cfg["ES_INDEX_TASK"], query,
                                         cfg["es_auth"], f"dup-history {label} {bi}/{len(batches)}"):
                for h in page:
                    if h["_id"] in seen_ids:
                        continue
                    seen_ids.add(h["_id"])
                    d  = h["_source"].get("Data") or {}
                    hh = d.get("householdId") or ""
                    if not hh:
                        continue
                    add = d.get("additionalDetails") or {}
                    user, created_ts = _user_and_ts(d)
                    given   = str(add.get("name") or "").strip()
                    family  = str(add.get("familyNameOfIndividual") or "").strip()
                    members = d.get("memberCount")
                    if members in (None, ""):
                        members = add.get("memberCount")
                    try:
                        members = int(float(members))
                    except (ValueError, TypeError):
                        members = None
                    area = str(add.get("administrativeArea") or "").strip()
                    fac_code = str((d.get("boundaryHierarchyCode") or {}).get("sppSfd", "") or "").strip()
                    history.append({
                        "date": _norm_date(d.get(date_field)), "ts": created_ts,
                        "user": user, "fac": None, "hh": hh,
                        "ident": _identity(f"{given} {family}", members, area or fac_code),
                        "head": f"{given} {family}".strip(), "family": family,
                        "members": members, "area": area,
                    })

    if families:
        _run_lookup("Data.additionalDetails.familyNameOfIndividual.keyword", families, "by-name")

    log.info(f"[analyze_itn] dup-history: {len(history):,} prior records "
             f"({len(families):,} family names probed)")
    return history


def _fetch_usernames(cfg, uuids):
    """
    Resolve acting-user UUIDs (task Data.createdBy) to usernames for the DUP
    sheets — display only; the matrix always compares raw UUIDs. Two batched
    terms lookups: project-staff (Data.userId -> userName/nameOfUser), then
    user-sync (Data.syncedUserId -> syncedUserName) for any still unresolved.
    Assumes staff userId / sync syncedUserId hold the same user UUID as task
    createdBy (the same key cdd_sync joins staff<->sync on). An unresolved
    UUID is displayed raw — a failed join must never hide who re-served a
    household.
    """
    result = {}
    if not uuids:
        return result

    def _lookup(index, id_field, name_fields, pending):
        found = {}
        batches = [pending[i:i + _BATCH] for i in range(0, len(pending), _BATCH)]
        for batch in batches:
            query = {
                "size": _BATCH,
                "query": {"terms": {f"{id_field}.keyword": batch}},
                "_source": [id_field] + name_fields,
            }
            for page in scroll_batches(cfg["es_url"], index, query, cfg["es_auth"],
                                         f"dup-usernames {index.split('-')[0]}"):
                for h in page:
                    d = h["_source"].get("Data") or {}
                    uid = str(d.get(id_field.split(".")[-1]) or "").strip()
                    if not uid or uid in found:
                        continue
                    for nf in name_fields:
                        name = str(d.get(nf.split(".")[-1]) or "").strip()
                        if name:
                            found[uid] = name
                            break
        return found

    pending = sorted(uuids)
    result.update(_lookup(cfg["ES_INDEX_STAFF"], "Data.userId",
                          ["Data.userName", "Data.nameOfUser"], pending))
    pending = sorted(set(pending) - set(result))
    if pending:
        result.update(_lookup(cfg["ES_INDEX_SYNC"], "Data.syncedUserId",
                              ["Data.syncedUserName"], pending))
    log.info(f"[analyze_itn] dup usernames: {len(result):,}/{len(uuids):,} UUIDs resolved")
    return result


def _classify_duplicates(events):
    """
    Classify every duplicate record in this run's window into the 2x2 matrix
    (same/different user x same/different day), each bucket attributed to the
    facility of the repeat record.

    Grouping: two records are the same physical household when they share the
    identity key (see _identity). householdId is display-only — it cannot
    connect re-registrations, where a re-served household gets a brand-new ID.

    Rules:
      - A group's original is its earliest record across window + history,
        ordered by (date, createdTime, householdId). Every other WINDOW record
        falls into exactly one bucket relative to that original.
      - History records (fac=None) are never counted as duplicates themselves —
        they were (or will be, via --cumulative) counted on their own day's run.
      - Missing user on either side, missing head name, or missing task date ->
        excluded, never guessed (a dateless record also cannot be a group's
        original — it would sort first and poison the day comparison). The
        caller warns when the excluded share is large.
      - Grouping is global, not per facility: a household re-served at a
        different facility is precisely the case this matrix exists to surface.
      - Ties within one date fall back to (scroll order, householdId) — chad
        docs carry no createdTime, so the su/du split of 3+ same-day
        mixed-user records is not guaranteed stable across reruns; totals are.
      - Two records both missing memberCount still match ("?" == "?") when
        name+village agree — consistent absence is treated as agreement.

    Returns (fac_buckets, detail_rows, dup_total):
      fac_buckets: {fac_code: {dup_su_sd: n, ...}} — facilities absent = all 0
      detail_rows: one dict per household with at least one classified repeat,
                   each repeat tagged with its bucket key ("type"); the
                   original (hh/date/user) plus all repeats ride on one row.
                   Matrix counts stay record-grain.
      dup_total:   total records classified (all four buckets)
    """
    groups = {}
    for e in events:
        # Dateless records are excluded here, not just at classification: ""
        # sorts before any date, so one would otherwise become the group's
        # "original" and force every real repeat into a different-day bucket.
        if e["ident"] and e["date"]:
            groups.setdefault(e["ident"], []).append(e)

    fac_buckets = {}
    detail_rows = []
    dup_total   = 0
    for grp in groups.values():
        if len(grp) < 2:
            continue
        grp.sort(key=lambda e: (e["date"], e["ts"], e["hh"]))
        orig = grp[0]
        repeats = []
        for e in grp[1:]:
            if e["fac"] is None:
                continue   # history record — counted on its own day, not here
            if not e["user"] or not orig["user"]:
                continue   # user unanswerable -> excluded, never guessed
            same_day = bool(e["date"]) and e["date"] == orig["date"]
            if e["user"] == orig["user"]:
                key = "dup_su_sd" if same_day else "dup_su_dd"
            else:
                key = "dup_du_sd" if same_day else "dup_du_dd"
            repeats.append({"date": e["date"], "user": e["user"],
                             "hh": e["hh"], "fac": e["fac"], "type": key})
            b = fac_buckets.setdefault(e["fac"], {k: 0 for k in _DUP_KEYS})
            b[key] += 1
            dup_total += 1
        if repeats:
            detail_rows.append({
                "head": orig["head"], "members": orig["members"],
                "area": orig.get("area", ""),
                "orig_hh": orig["hh"], "orig_date": orig["date"],
                "orig_user": orig["user"],
                "repeats": repeats,
            })
    return fac_buckets, detail_rows, dup_total


def _aggregate_to_lga(fac_rows):
    """
    Roll facility-level rows up to LGA grain — the grain that actually carries a
    real target (see module docstring). Keyed by lga_code (the DB join key), with
    the plain lga name kept for display.
    """
    lga_data = {}
    for r in fac_rows:
        code = r["lga_code"] or r["lga"] or "UNKNOWN"
        if code not in lga_data:
            lga_data[code] = dict(
                province=r["province"], lga=r["lga"] or code, lga_code=code,
                facilities=0,
                records=0, households_visited=0, nets_distributed=0, population_covered=0,
                missing_hh_head=0, missing_gps=0,
                manual_codes=0, scanned_codes=0, missing_codes=0,
            )
        d = lga_data[code]
        d["facilities"] += 1
        for k in ("records", "households_visited", "nets_distributed", "population_covered",
                  "missing_hh_head", "missing_gps", "manual_codes", "scanned_codes", "missing_codes"):
            d[k] += r[k]
        # Duplicate-matrix buckets: None means "not measured this run" (gate off
        # or enrichment failed — always uniform across rows) and must propagate,
        # never silently become a 0 that reads as "measured, none found".
        for k in _DUP_KEYS:
            rv = r.get(k)
            cur = d.get(k, 0)
            d[k] = None if (rv is None or cur is None) else cur + rv
    return list(lga_data.values())


# ── targets ────────────────────────────────────────────────────────────────────

def _load_targets_itn(cfg):
    """
    Reads a pre-built target CSV keyed by LGA CODE (chad.project_target via
    chad.project_address.boundary, boundarytype='LGA' — see module docstring for
    the exact query and the live ES cross-check proving code, not name, is the
    correct join key). Column names are matched case-insensitively against a set
    of known aliases so whatever export format is shared works without another
    round of edits:
      code column:       lga_code | district | boundary | LGA
      population target: population_target | individual_target | TargetPopulation
      household target:   household_target | TargetHouseholds
      net/ITN target:      net_target | product_target | TargetBednets
    Any column not found defaults to 0 (reported, not silently assumed correct).
    """
    from pipeline.core.drive import resolve_target_book
    csv_path = resolve_target_book(cfg)
    if not csv_path:
        log.error("[analyze_itn] no target book configured — all targets = 0, "
                  "so every coverage figure in this report will be 0% and meaningless — the report should not be shared until the target book is fixed")
        _degrade_run(TARGETS_ZERO + "no target book is configured for this ITN "
                     "campaign. Set target_file on the sheet row, or "
                     "DST_TARGET_FOLDER_ID for the deployment")
        return {}
    if csv_path.startswith("https://docs.google.com/spreadsheets/"):
        # same Sheets-URL form analyze.py accepts; without this the URL fails
        # os.path.exists and the ITN report publishes 0% coverage silently
        from pipeline.analyze import _read_target_sheet_url
        df = _read_target_sheet_url(csv_path)
        if df is None:
            log.error(f"[analyze_itn] could not read target sheet {csv_path} — ALL "
                      f"TARGETS = 0, so every coverage figure in this report is "
                      f"meaningless. Check target_file and DST_TARGET_FOLDER_ID")
            _degrade_run(TARGETS_ZERO + "the ITN target Google Sheet could not "
                         "be read: " + csv_path)
            return {}
    elif not os.path.exists(csv_path):
        log.error(f"[analyze_itn] target book not found: {csv_path} — all targets "
                  f"= 0, so every coverage figure in this report is meaningless "
                  f"— the report should not be shared until it is fixed")
        _degrade_run(TARGETS_ZERO + "the ITN target book was not found at "
                     + csv_path)
        return {}
    else:
        df = pd.read_csv(csv_path)
    cols_lower = {c.lower(): c for c in df.columns}

    def _find(*aliases):
        for a in aliases:
            if a.lower() in cols_lower:
                return cols_lower[a.lower()]
        return None

    code_col = _find("lga_code", "district", "boundary", "lga")
    pop_col  = _find("population_target", "individual_target", "targetpopulation")
    hh_col   = _find("household_target", "targethouseholds")
    net_col  = _find("net_target", "product_target", "targetbednets")

    if not code_col:
        log.warning(f"[analyze_itn] target_csv {csv_path} has no recognisable LGA-code "
                     f"column (tried lga_code/district/boundary/lga) — all targets = 0")
        return {}
    missing = [n for n, c in (("population", pop_col), ("household", hh_col), ("net", net_col)) if not c]
    if missing:
        log.warning(f"[analyze_itn] target_csv {csv_path} missing column(s) for: "
                     f"{', '.join(missing)} — those targets will read 0")

    def _num(row, col):
        if not col:
            return 0
        try:
            return int(float(row.get(col, 0) or 0))
        except (ValueError, TypeError):
            return 0

    tmap = {}
    for _, row in df.iterrows():
        code = str(row.get(code_col, "")).strip()
        if not code:
            continue
        tmap[code] = {
            "household_target":  _num(row, hh_col),
            "population_target": _num(row, pop_col),
            "net_target":        _num(row, net_col),
        }
    log.info(f"[analyze_itn] targets loaded: {len(tmap):,} LGAs")
    return tmap


# ── banding ────────────────────────────────────────────────────────────────────

def _cov_pct(numer, denom):
    return numer / denom * 100 if denom else 0.0


def _band(net_cov_pct, net_target, records):
    # Mirrors analyze.py's _band() priority order exactly. NOT REPORTED: zero
    # records = only the target-book zero rows (any ES-seen LGA has >= 1).
    if records == 0:
        return "NOT REPORTED"
    if records < 10:
        return "LOW ACTIVITY"
    if net_target == 0:
        return "NO TARGET"
    if net_cov_pct >= 95:
        return "HIGH"
    if net_cov_pct >= 70:
        return "MODERATE"
    return "LOW"


def _lga_name_from_code(code):
    """Readable LGA name from a boundary code: the segments after the last
    numeric one (ADMIN2_TC_01_01_01_ALAYE_GOYMO -> "ALAYE GOYMO")."""
    parts = str(code).split("_")
    last_num = max((i for i, p in enumerate(parts) if p.isdigit()), default=-1)
    return " ".join(parts[last_num + 1:]) or str(code)


def _finalize_lga_rows(lga_rows, target_map):
    """Rows are the UNION of LGAs that reported and the target book: a target-
    book LGA with no data becomes an explicit zero row (status NOT REPORTED),
    so the report's targets stay STANDARD irrespective of which LGAs synced —
    same rule as SPAQ's facility-level fix in analyze.py."""
    seen_codes = {r["lga_code"] for r in lga_rows}
    for code, tgt in target_map.items():
        if code in seen_codes:
            continue
        if not any(tgt.get(k, 0) > 0 for k in
                   ("household_target", "population_target", "net_target")):
            continue
        zero = dict(
            province="—", lga=_lga_name_from_code(code), lga_code=code,
            facilities=0,
            records=0, households_visited=0, nets_distributed=0, population_covered=0,
            missing_hh_head=0, missing_gps=0,
            manual_codes=0, scanned_codes=0, missing_codes=0,
        )
        for k in _DUP_KEYS:
            zero[k] = None
        lga_rows.append(zero)

    results = []
    for r in sorted(lga_rows, key=lambda x: (x["province"], x["lga"])):
        tgt = target_map.get(r["lga_code"],
                              {"household_target": 0, "population_target": 0, "net_target": 0})
        hh_cov  = _cov_pct(r["households_visited"], tgt["household_target"])
        pop_cov = _cov_pct(r["population_covered"], tgt["population_target"])
        net_cov = _cov_pct(r["nets_distributed"],   tgt["net_target"])
        # DQ signal: records vs distinct households — repeat/duplicate delivery
        # records per household (same spirit as SPAQ's duplicate tracking).
        dup_records = max(0, r["records"] - r["households_visited"])

        results.append({
            **r,
            "household_target":  tgt["household_target"],
            "population_target": tgt["population_target"],
            "net_target":         tgt["net_target"],
            "household_cov":      hh_cov,
            "population_cov":     pop_cov,
            "net_cov":            net_cov,
            "status":             _band(net_cov, tgt["net_target"], r["records"]),
            "dup_records":        dup_records,
        })
    return results


# ── Excel writing ──────────────────────────────────────────────────────────────

# Duplicate-matrix columns are appended at the END (after the existing DQ block)
# so every positional reader of the earlier columns — report_itn's row[:22] /
# row[:15] unpacks and _load_all_days_perf_itn's r[5]/r[8]/r[11] — keeps working
# against both pre-matrix and post-matrix files. Order must match _DUP_KEYS.
# Empty cell = matrix not measured that run (gate off/failed), distinct from 0.
_DUP_HEADERS = [
    "Dup Same User Same Day", "Dup Same User Diff Day",
    "Dup Diff User Same Day", "Dup Diff User Diff Day",
]

HEADERS = [
    "#", "Province", "LGA", "Facilities",
    "Target Households", "Households Visited", "HH Coverage %",
    "Target Population", "Population Covered", "Pop Coverage %",
    "Target ITNs", "Nets Distributed", "ITN Coverage %",
    "Status", "Records", "Duplicate Records",
    "Missing HH Head", "Missing GPS",
    "Manual Codes", "Scanned Codes", "% Scanned", "Missing Codes",
] + _DUP_HEADERS

# Facility-level detail tab has NO target/coverage/status columns — no facility-
# level target exists (see module docstring), so showing one would be fabricated.
FACILITY_HEADERS = [
    "#", "Province", "LGA", "Health Facility",
    "Records", "Duplicate Records", "Households Visited",
    "Nets Distributed", "Population Covered",
    "Missing HH Head", "Missing GPS",
    "Manual Codes", "Scanned Codes", "% Scanned", "Missing Codes",
] + _DUP_HEADERS


def _dm(v):
    """Duplicate-matrix cell value: None (not measured) -> empty cell, never 0."""
    return "" if v is None else v


# Conditional flags for the matrix columns (Status-column palette): different-
# user buckets red, same-user re-visit orange, same-user same-day (sync-retry
# noise) unflagged. Applied only to non-zero values — a zero is not an alert.
_DUP_FLAG_COLOR = {
    "Dup Same User Diff Day": "E06000",
    "Dup Diff User Same Day": "CC0000",
    "Dup Diff User Diff Day": "CC0000",
}


def _row_values(r, idx):
    total_codes = r["manual_codes"] + r["scanned_codes"]
    pct_scanned = f"{r['scanned_codes']/total_codes*100:.1f}%" if total_codes else "N/A"
    return [
        idx, r["province"], r["lga"], r["facilities"],
        r["household_target"], r["households_visited"], f"{r['household_cov']:.1f}%",
        r["population_target"], r["population_covered"], f"{r['population_cov']:.1f}%",
        r["net_target"], r["nets_distributed"], f"{r['net_cov']:.1f}%",
        r["status"], r["records"], r["dup_records"],
        r["missing_hh_head"], r["missing_gps"],
        r["manual_codes"], r["scanned_codes"], pct_scanned, r["missing_codes"],
    ] + [_dm(r.get(k)) for k in _DUP_KEYS]


def _facility_row_values(r, idx):
    total_codes = r["manual_codes"] + r["scanned_codes"]
    pct_scanned = f"{r['scanned_codes']/total_codes*100:.1f}%" if total_codes else "N/A"
    return [
        idx, r["province"], r["lga"], r["facility_name"],
        r["records"], r.get("dup_records", 0), r["households_visited"],
        r["nets_distributed"], r["population_covered"],
        r["missing_hh_head"], r["missing_gps"],
        r["manual_codes"], r["scanned_codes"], pct_scanned, r["missing_codes"],
    ] + [_dm(r.get(k)) for k in _DUP_KEYS]


def _totals_row(rows):
    def s(k): return sum(r[k] for r in rows)

    def sn(k):
        # None-propagating sum for the duplicate-matrix buckets (None = not
        # measured this run, uniform across rows — see _DUP_KEYS).
        vals = [r.get(k) for r in rows]
        return None if any(v is None for v in vals) else sum(vals)

    hh_t, pop_t, net_t = s("household_target"), s("population_target"), s("net_target")
    hh_v, pop_v, net_v = s("households_visited"), s("population_covered"), s("nets_distributed")
    return {
        "province": "", "lga": "GRAND TOTAL", "facilities": s("facilities"),
        "household_target": hh_t, "households_visited": hh_v,
        "household_cov": _cov_pct(hh_v, hh_t),
        "population_target": pop_t, "population_covered": pop_v,
        "population_cov": _cov_pct(pop_v, pop_t),
        "net_target": net_t, "nets_distributed": net_v,
        "net_cov": _cov_pct(net_v, net_t),
        "status": "", "records": s("records"), "dup_records": s("dup_records"),
        "missing_hh_head": s("missing_hh_head"), "missing_gps": s("missing_gps"),
        "manual_codes": s("manual_codes"), "scanned_codes": s("scanned_codes"),
        "missing_codes": s("missing_codes"),
        **{k: sn(k) for k in _DUP_KEYS},
    }


def _write_tab(ws, rows, banner_text):
    ncols = len(HEADERS)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    banner = ws["A1"]
    banner.value = banner_text
    banner.fill  = BANNER_FILL
    banner.font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        style_cell(cell, fill=HDR_FILL, bold=True, align="center", color="FFFFFF")

    status_col = HEADERS.index("Status") + 1
    dup_flag_cols = {HEADERS.index(h) + 1: c for h, c in _DUP_FLAG_COLOR.items()}
    for ri, r in enumerate(rows, 1):
        vals = _row_values(r, ri)
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            style_cell(cell, fill=WHITE_FILL, align="center")
            if ci == status_col:
                flag_col = FLAG_COLOR.get(str(val), "000000")
                cell.font = Font(bold=True, color=flag_col, size=9, name="Calibri")
            elif ci in dup_flag_cols and isinstance(val, int) and val > 0:
                cell.font = Font(bold=True, color=dup_flag_cols[ci], size=9, name="Calibri")

    if rows:
        tot = _totals_row(rows)
        tot_row = len(rows) + 3
        tot_vals = _row_values(tot, "")
        for ci, val in enumerate(tot_vals, 1):
            cell = ws.cell(row=tot_row, column=ci, value=val)
            style_cell(cell, fill=TOTAL_FILL, bold=True, align="center")

    ws.freeze_panes = "E3"
    ws.auto_filter.ref = f"A2:{last_col}2"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    for ci in range(5, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _write_facility_tab(ws, rows, banner_text):
    ncols = len(FACILITY_HEADERS)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    banner = ws["A1"]
    banner.value = banner_text
    banner.fill  = BANNER_FILL
    banner.font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for ci, h in enumerate(FACILITY_HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        style_cell(cell, fill=HDR_FILL, bold=True, align="center", color="FFFFFF")

    dup_flag_cols = {FACILITY_HEADERS.index(h) + 1: c for h, c in _DUP_FLAG_COLOR.items()}
    for ri, r in enumerate(sorted(rows, key=lambda x: x["records"]), 1):
        vals = _facility_row_values(r, ri)
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            style_cell(cell, fill=WHITE_FILL, align="center")
            if ci in dup_flag_cols and isinstance(val, int) and val > 0:
                cell.font = Font(bold=True, color=dup_flag_cols[ci], size=9, name="Calibri")

    ws.freeze_panes = "E3"
    ws.auto_filter.ref = f"A2:{last_col}2"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 32
    for ci in range(5, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14


# "CRID" = household clientReferenceId: the task doc's Data.householdId holds
# the CLIENT reference (verified against the household index — the server id
# is the idgen "H-..." form), so these columns are the correct join key back
# to the household index / DB (household.clientreferenceid).
# Location block follows the boundary hierarchy Province > LGA > Facility >
# Village; Facility is the first repeat's distribution point.
_DUP_DETAIL_HEADERS = [
    "#", "Province", "LGA", "Facility", "Village", "Head Name", "Members",
    "Original Household CRID", "Original Date", "Original User",
    "Repeats", "Repeat Dates", "Repeat Users", "Repeat Household CRIDs",
]

# One sheet per duplicate type, ordered most severe first (a different user
# on a later day = double distribution, the case that costs nets). A mixed
# household appears on each sheet where it has a repeat of that type, listing
# only those repeats there. Sheet names must start with "DUP" — report_itn
# strips every DUP* sheet from the partner Excel by that prefix.
_DUP_SHEET_ORDER = ("dup_du_dd", "dup_du_sd", "dup_su_dd", "dup_su_sd")
_DUP_SHEET_NAME = {
    "dup_du_dd": "DUP DIFF USER DIFF DAY",
    "dup_du_sd": "DUP DIFF USER SAME DAY",
    "dup_su_dd": "DUP SAME USER DIFF DAY",
    "dup_su_sd": "DUP SAME USER SAME DAY",
}


def _split_dup_detail_by_type(detail_rows):
    """
    {bucket key: household rows for that sheet}. A household lands on every
    sheet where it has >=1 repeat of that type, carrying ONLY that type's
    repeats there — so each sheet is self-contained and needs no type column.
    """
    by_type = {}
    for r in detail_rows:
        for t in _DUP_SHEET_ORDER:
            reps = [p for p in r["repeats"] if p["type"] == t]
            if reps:
                by_type.setdefault(t, []).append({**r, "repeats": reps})
    return by_type


def _write_dup_detail_tab(ws, detail_rows, fac_meta, banner_text, uname_map=None):
    """
    One DUP sheet for one duplicate type (the sheet name IS the type — see
    _DUP_SHEET_NAME): one row per household with that kind of repeat, all its
    repeats joined "; "-separated and position-aligned across the repeat
    columns. Uncapped — each sheet is the complete census for its type.
    User columns show usernames resolved via _fetch_usernames; an unresolved
    UUID is shown raw rather than blanked. Internal-only: every DUP* sheet is
    dropped whole from the partner Excel (head names, household CRIDs and
    users must never leave the internal report).
    Values are device-entered free text, sanitised here: control characters
    are stripped (openpyxl raises IllegalCharacterError on them), "="-prefixed
    strings are forced to plain text (openpyxl would store them as live
    formulas), and over-long joined cells are truncated with a visible
    "(+N more)" tail (Excel's cell limit is 32,767 chars — truncation must
    never be silent).
    """
    ncols = len(_DUP_DETAIL_HEADERS)
    last_col = get_column_letter(ncols)

    def _clean(v):
        return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v

    def _join(parts):
        parts = [str(p) for p in parts]
        s = "; ".join(parts)
        if len(s) > 2000:
            keep, total = [], 0
            for p in parts:
                if total + len(p) > 1900:
                    break
                keep.append(p)
                total += len(p) + 2
            s = "; ".join(keep) + f" (+{len(parts) - len(keep)} more)"
        return s

    ws.merge_cells(f"A1:{last_col}1")
    banner = ws["A1"]
    banner.value = banner_text
    banner.fill  = BANNER_FILL
    banner.font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for ci, h in enumerate(_DUP_DETAIL_HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        style_cell(cell, fill=HDR_FILL, bold=True, align="center", color="FFFFFF")

    def _meta(fc):
        m = fac_meta.get(fc) or {}
        return m.get("province", ""), m.get("lga", ""), m.get("facility_name", fc)

    def _first_fac(r):
        return r["repeats"][0]["fac"] if r.get("repeats") else ""

    uname_map = uname_map or {}

    def _uname(u):
        return uname_map.get(u, u)

    shown = sorted(detail_rows,
                   key=lambda r: (_meta(_first_fac(r))[1], r.get("head", "")))
    for ri, r in enumerate(shown, 1):
        province, lga, fac_name = _meta(_first_fac(r))
        members = r.get("members")
        reps = r.get("repeats") or []
        vals = [ri, province, lga, fac_name, r.get("area", ""),
                r.get("head", ""), "" if members is None else members,
                r.get("orig_hh", ""), r["orig_date"], _uname(r["orig_user"]),
                len(reps),
                _join(p["date"] for p in reps),
                _join(_uname(p["user"]) for p in reps),
                _join(p["hh"] for p in reps)]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri + 2, column=ci, value=_clean(val))
            if isinstance(val, str) and val.startswith("="):
                cell.data_type = "s"
            style_cell(cell, fill=WHITE_FILL, align="center")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_col}2"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 32
    for ci in range(5, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22


BANDS = ["LOW", "MODERATE", "HIGH", "NO TARGET", "LOW ACTIVITY", "NOT REPORTED"]


# ── public entry point ─────────────────────────────────────────────────────────

def run(cfg):
    log.info(f"[analyze_itn] {cfg['state_name']} — streaming ITN task docs ...")

    fac_rows, dup_events = _fetch_facility_rows(cfg)
    # Duplicate records per facility, carried into the facility-detail tab (not
    # part of the LGA-level accumulator dict, so computed once here).
    for r in fac_rows:
        r["dup_records"] = max(0, r["records"] - r["households_visited"])

    # Duplicate-distribution matrix — opt-in (dup_matrix=TRUE on the sheet row)
    # and FAIL-SOFT by design, same doctrine as run.py's non-fatal cdd_sync step:
    # a DQ enrichment must never block the coverage report. Any failure here
    # leaves the matrix columns as None (empty cells, "not measured"), logs a
    # warning, and the rest of this run proceeds byte-identically to before.
    dup_detail = []
    if cfg.get("dup_matrix"):
        try:
            # Coverage guards: name, user and date are all required to classify
            # a duplicate — records missing any are excluded, so a largely-blank
            # field must be warned about, never published as quiet zeros.
            no_ident = sum(1 for e in dup_events if not e["ident"])
            no_user  = sum(1 for e in dup_events if not e["user"])
            no_date  = sum(1 for e in dup_events if not e["date"])
            if dup_events and no_ident / len(dup_events) > 0.5:
                log.warning(f"[analyze_itn] dup matrix: {no_ident:,}/{len(dup_events):,} records "
                            f"have no head name — identity matching will undercount duplicates")
            if dup_events and no_user / len(dup_events) > 0.5:
                log.warning(f"[analyze_itn] dup matrix: {no_user:,}/{len(dup_events):,} records "
                            f"have no user (createdBy) — such duplicates are excluded, "
                            f"matrix will undercount")
            if no_date:
                log.warning(f"[analyze_itn] dup matrix: {no_date:,}/{len(dup_events):,} records "
                            f"have no task date — excluded from the matrix (day dimension "
                            f"unanswerable), matrix will undercount")
            # Cumulative runs already scroll the whole campaign window — no
            # history lookup needed.
            history = [] if cfg.get("cumulative") else _fetch_dup_history(cfg, dup_events)
            fac_buckets, dup_detail, dup_total = _classify_duplicates(dup_events + history)
            zero = {k: 0 for k in _DUP_KEYS}
            for r in fac_rows:
                r.update(fac_buckets.get(r["facility_code"], zero))
            log.info(f"[analyze_itn] dup matrix: {dup_total:,} duplicate records classified "
                     f"across {len(dup_detail):,} households")
            # Username join has its own fail-soft: a broken staff/sync lookup
            # falls back to raw UUIDs, never blocks the matrix.
            try:
                uuids = {d["orig_user"] for d in dup_detail} | \
                        {p["user"] for d in dup_detail for p in d["repeats"]}
                dup_unames = _fetch_usernames(cfg, {u for u in uuids if u})
            except Exception as e:
                log.warning(f"[analyze_itn] dup username join failed (non-fatal — UUIDs shown): {e}")
                dup_unames = {}
        except Exception as e:
            log.warning(f"[analyze_itn] dup matrix failed (non-fatal — report proceeds without it): {e}")
            dup_detail, dup_unames = [], {}
            for r in fac_rows:
                r.update({k: None for k in _DUP_KEYS})
    else:
        dup_unames = {}
        for r in fac_rows:
            r.update({k: None for k in _DUP_KEYS})

    lga_rows_raw = _aggregate_to_lga(fac_rows)
    target_map   = _load_targets_itn(cfg)

    # Daily runs compare today's activity against a DAILY target, not the full
    # campaign target — same "Total Campaign Target = Daily Target x Campaign
    # Days" relationship analyze.py's own daily mode uses. The target CSV holds
    # the full campaign-length figure, so divide it down here for a daily run;
    # cumulative runs (--cumulative) use the full figure as-is.
    if not cfg.get("cumulative"):
        campaign_start = cfg.get("campaign_start")
        campaign_end   = cfg.get("campaign_end")
        campaign_days  = (campaign_end - campaign_start).days + 1 if campaign_start and campaign_end else None
        if campaign_days:
            target_map = {
                code: {k: round(v / campaign_days) for k, v in t.items()}
                for code, t in target_map.items()
            }

    rows = _finalize_lga_rows(lga_rows_raw, target_map)

    g = _totals_row(rows) if rows else None
    _bdays = None
    if cfg.get("campaign_start") and cfg.get("campaign_end"):
        _bdays = (cfg["campaign_end"] - cfg["campaign_start"]).days + 1
    if g and not cfg.get("cumulative") and _bdays:
        banner_text = (
            f"Daily Targets (= Campaign Target \u00f7 {_bdays} campaign days) — "
            f"Households: {g['household_target']:,}  |  "
            f"Population: {g['population_target']:,}  |  "
            f"ITNs: {g['net_target']:,}  "
            f"(all targets/coverage in this sheet are DAILY)"
        )
    elif g:
        banner_text = (
            f"Overall Campaign Targets — Households: {g['household_target']:,}  |  "
            f"Population: {g['population_target']:,}  |  "
            f"ITNs: {g['net_target']:,}"
        )
    else:
        banner_text = "No data"

    wb = Workbook()
    wb.remove(wb.active)

    ws_all = wb.create_sheet("ALL LGAS")
    _write_tab(ws_all, rows, banner_text)

    for band in BANDS:
        band_rows = [r for r in rows if r["status"] == band]
        ws = wb.create_sheet(band)
        _write_tab(ws, band_rows, banner_text)

    ws_fac = wb.create_sheet("FACILITY DETAIL")
    _write_facility_tab(ws_fac, fac_rows, banner_text)

    # Household-level duplicate trace — one sheet per type (severity order),
    # each written only when it has rows (an empty tab would imply "checked,
    # clean" on runs where the matrix was never measured at all). Guarded:
    # this writes device-entered free text and runs after the fail-soft dup
    # block, so a writer failure must cost the DUP sheets only, never the
    # legacy tabs or the workbook save.
    if dup_detail:
        try:
            fac_meta = {r["facility_code"]: r for r in fac_rows}
            by_type = _split_dup_detail_by_type(dup_detail)
            for t in _DUP_SHEET_ORDER:
                if by_type.get(t):
                    ws_dup = wb.create_sheet(_DUP_SHEET_NAME[t])
                    _write_dup_detail_tab(ws_dup, by_type[t], fac_meta, banner_text, dup_unames)
        except Exception as e:
            log.warning(f"[analyze_itn] DUP sheets failed (non-fatal — legacy tabs kept): {e}")
            for t in _DUP_SHEET_ORDER:
                if _DUP_SHEET_NAME[t] in wb.sheetnames:
                    wb.remove(wb[_DUP_SHEET_NAME[t]])

    out = cfg["perf_xlsx"]
    wb.save(out)
    log.info(f"[analyze_itn] saved -> {out}  ({len(rows)} LGAs, {len(fac_rows)} facilities)")

    # Day-indexed history snapshot, mirroring SPAQ's performance_day{N}.xlsx so
    # report_itn can build a real cumulative series. cfg["perf_xlsx"] cannot be
    # used: it is named from cfg["DAY"], which config.py clamps to campaign_days,
    # so every day would overwrite one file. Keyed on the REAL elapsed day.
    # Never in cumulative mode - the wb then holds campaign-wide figures and
    # would overwrite that day's genuine daily snapshot.
    if cfg.get("campaign_start") and cfg.get("extract_date") and not cfg.get("cumulative"):
        elapsed_day = (cfg["extract_date"] - cfg["campaign_start"]).days + 1
        hist_dir = os.path.join(os.path.dirname(out), "itn_history")
        os.makedirs(hist_dir, exist_ok=True)
        hist_path = os.path.join(hist_dir, f"performance_day{elapsed_day}.xlsx")
        wb.save(hist_path)
        log.info(f"[analyze_itn] history snapshot -> {hist_path} (elapsed day {elapsed_day})")

    return out
