"""openpyxl styling primitives shared by the performance and CDD-sync workbooks."""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_thin = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

BANNER_FILL = PatternFill("solid", fgColor="17365D")
HDR_FILL    = PatternFill("solid", fgColor="1A6496")
TOTAL_FILL  = PatternFill("solid", fgColor="EEEEEE")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

SYNC_HDR_FILL   = PatternFill("solid", fgColor="003366")
SYNC_NEVER_FILL = PatternFill("solid", fgColor="FFD7D7")
SYNC_LOW_FILL   = PatternFill("solid", fgColor="FFE0B3")
SYNC_TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")

FLAG_COLOR = {
    "HIGH":         "1A7A1A",
    "MODERATE":     "E06000",
    "LOW":          "CC0000",
    "NO TARGET":    "888888",
    "LOW ACTIVITY": "888888",
    "NOT REPORTED": "888888",
}


def style_cell(cell, fill=None, bold=False, color=None, align="center", size=9):
    cell.border = BORDER
    if fill:
        cell.fill = fill
    font_kwargs = {"bold": bold, "size": size, "name": "Calibri"}
    if color:
        font_kwargs["color"] = color
    cell.font = Font(**font_kwargs)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


def style_sync_cell(cell, fill=None, bold=False, color=None):
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill
    cell.font = Font(bold=bold, size=9, name="Calibri",
                     color=color if color else "000000")
