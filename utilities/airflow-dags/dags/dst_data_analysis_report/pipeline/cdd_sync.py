"""
cdd_sync.py — ES staff index scroll + composite agg → CDD sync Excel

Two variants:
  is_admin_console=TRUE  (Nigeria): filter staff by campaignNumber, sync by syncedUserName
  is_admin_console=FALSE (Chad/Togo): filter staff by projectTypeId, sync by syncedUserId
"""
import logging
import os
from collections import defaultdict
from datetime import timedelta, datetime, timezone

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from dst_data_analysis_report.pipeline.core.checkpoint import load_checkpoint, save_checkpoint
from dst_data_analysis_report.pipeline.core.es import composite_agg, scroll_all
from dst_data_analysis_report.pipeline.core.excel import (
    SYNC_HDR_FILL, SYNC_LOW_FILL, SYNC_NEVER_FILL, SYNC_TOTAL_FILL, style_sync_cell,
)

log = logging.getLogger(__name__)

# The CDD role is DEPLOYMENT config, not a constant. It was hardcoded in nine
# separate term filters below, so standing up Togo meant hand-editing all nine
# to COMMUNITY_DISTRIBUTOR on every pipeline sync — a swap that was lost and
# re-applied more than once. deployment_env.py already DOCUMENTED a CDD_ROLE
# group env key; nothing ever read it. Now it does.
#
# Deliberately separate from the ITN role (CDD_ROLE_ITN in cdd_sync_itn.py):
# one deployment runs both campaign types at once — Bauchi has had SMC and ITN
# live in the same window — so a single shared key would break one of them.
def _cdd_role():
    return os.getenv("CDD_ROLE", "DISTRIBUTOR").strip() or "DISTRIBUTOR"




def _epoch_ms_to_date(val):
    if isinstance(val, int):
        return datetime.fromtimestamp(val / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
    return str(val)[:10]


# ── Variant A: admin console = TRUE (Nigeria, campaignNumber) ──────────────────

def _load_staff_by_campaign(cfg):
    query = {
        "size": 5000,
        "_source": [
            "Data.userId", "Data.userName", "Data.nameOfUser",
            "Data.boundaryHierarchy.healthFacility",
            "Data.boundaryHierarchy.lga",
        ],
        "query": {"bool": {"must": [
            {"term": {"Data.campaignNumber.keyword": cfg["campaign_number"]}},
            {"term": {"Data.role.keyword": _cdd_role()}},
        ]}},
    }
    hits = scroll_all(cfg["es_url"], cfg["ES_INDEX_STAFF"], query,
                       cfg["es_auth"], "staff (admin)")
    cdds       = {}   # lower(uname) -> info
    uname_list = []
    for h in hits:
        d     = h["_source"]["Data"]
        uname = (d.get("userName", "") or d.get("nameOfUser", "")).strip()
        key   = uname.lower()
        if not key or key in cdds:
            continue
        bh = d.get("boundaryHierarchy") or {}
        cdds[key] = {
            "user_id":  d.get("userId", "").strip(),
            "username": uname,
            "facility": bh.get("healthFacility", ""),
            "lga":      bh.get("lga", ""),
        }
        uname_list.append(uname)
    log.info(f"  staff (admin): {len(cdds):,} unique CDDs")
    return cdds, uname_list


def _load_sync_dates_by_username(cfg, uname_list):
    must = [
        {"terms": {"Data.taskDates":              cfg["CAMPAIGN_DATES"]}},
        {"terms": {"Data.syncedUserName.keyword": uname_list}},
        {"term": {"Data.role.keyword": _cdd_role()}},
        {"term":  {"Data.campaignNumber.keyword": cfg["campaign_number"]}},
    ]
    sources = [
        {"uname": {"terms": {"field": "Data.syncedUserName.keyword"}}},
        {"date":  {"terms": {"field": "Data.taskDates"}}},
    ]
    buckets = composite_agg(cfg["es_url"], cfg["ES_INDEX_SYNC"], must, sources,
                              cfg["es_auth"])
    log.info(f"  sync agg (admin): {len(buckets):,} (uname, date) pairs")
    user_dates = defaultdict(set)
    for b in buckets:
        uname = b["key"]["uname"].lower()
        dt    = _epoch_ms_to_date(b["key"]["date"])
        if dt in cfg["CAMPAIGN_DATES"]:
            user_dates[uname].add(dt)
    return user_dates


# ── Variant B: admin console = FALSE (Chad/Togo, projectTypeId) ───────────────

def _load_staff_by_project_type(cfg):
    query = {
        "size": 5000,
        "_source": [
            "Data.userId", "Data.userName", "Data.nameOfUser",
            "Data.boundaryHierarchy",
        ],
        "query": {"bool": {"must": [
            {"term": {"Data.projectTypeId.keyword": cfg["project_type_id"]}},
            {"term": {"Data.role.keyword": _cdd_role()}},
            {"term": {"Data.isDeleted":             False}},
        ]}},
    }
    hits = scroll_all(cfg["es_url"], cfg["ES_INDEX_STAFF"], query,
                       cfg["es_auth"], "staff (project)")
    cdds     = {}   # lower(uid) -> info
    uid_list = []
    for h in hits:
        d   = h["_source"]["Data"]
        uid = (d.get("userId", "") or "").strip().lower()
        if not uid or uid in cdds:
            continue
        bh = d.get("boundaryHierarchy") or {}
        cdds[uid] = {
            "user_id":    d.get("userId", "").strip(),
            "username":   (d.get("userName", "") or d.get("nameOfUser", "")).strip(),
            "province":   bh.get("province",    ""),
            "district":   bh.get("district",    ""),
            "health_area":bh.get("healthArea",  ""),
            "facility":   bh.get("healthFacility", "") or bh.get("HEALTHFACILITY", ""),
            "lga":        bh.get("lga", "") or bh.get("district", ""),
        }
        uid_list.append(d.get("userId", "").strip())
    log.info(f"  staff (project): {len(cdds):,} unique CDDs")
    return cdds, uid_list


def _load_sync_dates_by_user_id(cfg, uid_list):
    must = [
        {"terms": {"Data.taskDates":            cfg["CAMPAIGN_DATES"]}},
        {"terms": {"Data.syncedUserId.keyword": uid_list}},
    ]
    sources = [
        {"uid":  {"terms": {"field": "Data.syncedUserId.keyword"}}},
        {"date": {"terms": {"field": "Data.taskDates"}}},
    ]
    buckets = composite_agg(cfg["es_url"], cfg["ES_INDEX_SYNC"], must, sources,
                              cfg["es_auth"])
    log.info(f"  sync agg (project): {len(buckets):,} (uid, date) pairs")
    user_dates = defaultdict(set)
    for b in buckets:
        uid = b["key"]["uid"].lower()
        dt  = _epoch_ms_to_date(b["key"]["date"])
        if dt in cfg["CAMPAIGN_DATES"]:
            user_dates[uid].add(dt)
    return user_dates


# ── time-cutoff sync count ────────────────────────────────────────────────────

def _count_synced_by_cutoff(cfg, cutoff_hour, cutoff_min=0):
    """
    Count unique CDDs who synced TODAY before cutoff_hour:cutoff_min UTC.
    Uses Data.createdTime (epoch ms) as the server receipt timestamp.
    Returns int count, or None if query fails.
    """
    today = cfg["extract_date"].isoformat()
    cutoff_dt = datetime.strptime(
        f"{today}T{cutoff_hour:02d}:{cutoff_min:02d}:00", "%Y-%m-%dT%H:%M:%S"
    )
    cutoff_ms = int(cutoff_dt.replace(tzinfo=timezone.utc).timestamp() * 1000)

    if cfg.get("tenant") == "ch":
        # Chad: the user-sync-index carries campaignNumber (NOT projectTypeId),
        # and the CDD role is DISTRIBUTOR. Scope on campaignNumber so the cutoff
        # is non-zero (projectTypeId doesn't exist on this index).
        must_filter = [
            {"term":  {"Data.campaignNumber.keyword": cfg["campaign_number"]}},
            {"term":  {"Data.role.keyword": _cdd_role()}},
            {"terms": {"Data.taskDates": [today]}},
            {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
        ]
        agg_field = "Data.syncedUserId.keyword"
    elif cfg["is_admin_console"]:
        must_filter = [
            {"term":  {"Data.campaignNumber.keyword": cfg["campaign_number"]}},
            {"term": {"Data.role.keyword": _cdd_role()}},
            {"terms": {"Data.taskDates": [today]}},
            {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
        ]
        agg_field = "Data.syncedUserName.keyword"
    else:
        must_filter = [
            {"term":  {"Data.projectTypeId.keyword": cfg["project_type_id"]}},
            {"term": {"Data.role.keyword": _cdd_role()}},
            {"terms": {"Data.taskDates": [today]}},
            {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
        ]
        agg_field = "Data.syncedUserId.keyword"

    query = {
        "size": 0,
        "query": {"bool": {"filter": must_filter}},
        "aggs": {"unique_synced": {"cardinality": {"field": agg_field, "precision_threshold": 3000}}},
    }
    try:
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_SYNC']}/_search",
            json=query, auth=cfg["es_auth"], verify=False, timeout=30,
        )
        r.raise_for_status()
        count = r.json()["aggregations"]["unique_synced"]["value"]
        log.info(f"  synced by {cutoff_hour:02d}:{cutoff_min:02d} UTC: {count:,}")
        return count
    except Exception as e:
        log.warning(f"  time-cutoff query failed (non-fatal): {e}")
        return None


def _get_synced_keys_by_cutoff(cfg, cutoff_hour=17, cutoff_min=30):
    """
    Return set of CDD keys (lowercased) who synced TODAY before cutoff_hour:cutoff_min UTC.
    Admin console: keys are syncedUserName.lower().
    Project: keys are syncedUserId.lower().
    Returns set, or None if query fails.
    """
    today = cfg["extract_date"].isoformat()
    cutoff_dt = datetime.strptime(
        f"{today}T{cutoff_hour:02d}:{cutoff_min:02d}:00", "%Y-%m-%dT%H:%M:%S"
    )
    cutoff_ms = int(cutoff_dt.replace(tzinfo=timezone.utc).timestamp() * 1000)

    if cfg.get("tenant") == "ch":
        # Chad: sync index has campaignNumber (not projectTypeId); role DISTRIBUTOR.
        must_filter = [
            {"term":  {"Data.campaignNumber.keyword": cfg["campaign_number"]}},
            {"term":  {"Data.role.keyword": _cdd_role()}},
            {"terms": {"Data.taskDates": [today]}},
            {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
        ]
        agg_sources = [{"key": {"terms": {"field": "Data.syncedUserId.keyword"}}}]
    elif cfg["is_admin_console"]:
        must_filter = [
            {"term":  {"Data.campaignNumber.keyword": cfg["campaign_number"]}},
            {"term": {"Data.role.keyword": _cdd_role()}},
            {"terms": {"Data.taskDates": [today]}},
            {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
        ]
        agg_sources = [{"key": {"terms": {"field": "Data.syncedUserName.keyword"}}}]
    else:
        must_filter = [
            {"term":  {"Data.projectTypeId.keyword": cfg["project_type_id"]}},
            {"term": {"Data.role.keyword": _cdd_role()}},
            {"terms": {"Data.taskDates": [today]}},
            {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
        ]
        agg_sources = [{"key": {"terms": {"field": "Data.syncedUserId.keyword"}}}]

    try:
        buckets = composite_agg(
            cfg["es_url"], cfg["ES_INDEX_SYNC"], must_filter, agg_sources, cfg["es_auth"]
        )
        synced = {b["key"]["key"].lower() for b in buckets}
        log.info(f"  synced by {cutoff_hour:02d}:{cutoff_min:02d} UTC: {len(synced):,} CDDs")
        return synced
    except Exception as e:
        log.warning(f"  cutoff key query failed (non-fatal): {e}")
        return None


# ── status + build rows ────────────────────────────────────────────────────────

def _sync_status_band(n, day):
    if n == day:      return "HIGH"
    elif n >= 3:      return "MODERATE"
    elif n >= 1:      return "LOW"
    else:             return "NEVER SYNCED"


def _build_cdd_rows(cdds, user_dates, cfg, is_admin):
    day_labels = {}
    start = cfg["campaign_start"]
    for i, dt in enumerate(cfg["CAMPAIGN_DATES"]):
        d = start + timedelta(days=i)
        day_labels[dt] = f"Day {i+1} ({d.day} {d.strftime('%b')})"

    all_rows = []
    for key, info in cdds.items():
        dates_synced = user_dates.get(key, set())
        n_days       = len(dates_synced)
        rec = {
            "LGA":             info.get("lga", "") or info.get("province", ""),
            "Health Facility": info.get("facility", "") or info.get("health_area", ""),
            "Username":        info["username"],
            "User ID":         info["user_id"],
            "Days Synced":     n_days,
        }
        for dt, col in day_labels.items():
            rec[col] = "Y" if dt in dates_synced else "N"
        rec["Status"]     = _sync_status_band(n_days, cfg["DAY"])
        rec["Sync Dates"] = ", ".join(sorted(dates_synced))
        all_rows.append(rec)

    return all_rows, day_labels


# ── Excel writer ───────────────────────────────────────────────────────────────

def _write_dataframe_tab(ws, df, hdr_fill=SYNC_HDR_FILL):
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        style_sync_cell(cell, fill=hdr_fill, bold=True, color="FFFFFF")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        status = getattr(row, "Status", None)
        row_fill = (
            SYNC_NEVER_FILL if status == "NEVER SYNCED"
            else SYNC_LOW_FILL if status == "LOW"
            else None
        )
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            style_sync_cell(cell, fill=row_fill)
    # auto-width
    for ci, col in enumerate(df.columns, 1):
        lengths = [len(str(col))] + [len(str(v)) for v in df.iloc[:, ci - 1] if v is not None]
        max_len = max(lengths) if lengths else 10
        ws.column_dimensions[
            get_column_letter(ci)
        ].width = min(max_len + 2, 30)


def _write_summary_tab(ws, all_rows, cfg):
    lga_stats = defaultdict(lambda: {"total": 0, "HIGH": 0, "MODERATE": 0, "LOW": 0, "NEVER SYNCED": 0})
    for rec in all_rows:
        lg = rec["LGA"] or "Unknown"
        lga_stats[lg]["total"] += 1
        lga_stats[lg][rec["Status"]] += 1

    cols = ["#", "LGA", "Total CDDs", "HIGH", "MODERATE", "LOW", "NEVER SYNCED", "% Never Synced"]
    for ci, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        style_sync_cell(cell, fill=SYNC_HDR_FILL, bold=True, color="FFFFFF")

    for ri, (lga, s) in enumerate(sorted(lga_stats.items()), 2):
        pct = f"{s['NEVER SYNCED']/s['total']*100:.1f}%" if s["total"] else "-"
        vals = [ri - 1, lga, s["total"], s["HIGH"], s["MODERATE"], s["LOW"], s["NEVER SYNCED"], pct]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            style_sync_cell(cell)

    # grand total
    tr = len(lga_stats) + 2
    totals = [
        "", "GRAND TOTAL",
        sum(s["total"] for s in lga_stats.values()),
        sum(s["HIGH"] for s in lga_stats.values()),
        sum(s["MODERATE"] for s in lga_stats.values()),
        sum(s["LOW"] for s in lga_stats.values()),
        sum(s["NEVER SYNCED"] for s in lga_stats.values()),
        "",
    ]
    for ci, val in enumerate(totals, 1):
        cell = ws.cell(row=tr, column=ci, value=val)
        style_sync_cell(cell, fill=SYNC_TOTAL_FILL, bold=True)

    return lga_stats


# ── public entry point ─────────────────────────────────────────────────────────

def collect(cfg):
    """Stage 1 (all ES I/O): staff roster, per-day sync aggregation, today's cutoff counts.

    Returns a JSON-safe dict, or None when the campaign is not queryable
    (missing identifiers or an empty roster).
    """
    is_admin = cfg["is_admin_console"]

    if is_admin:
        if not cfg["campaign_number"]:
            log.warning("[cdd_sync] is_admin_console=TRUE but campaign_number is empty — skipping")
            return None
        cdds, key_list = _load_staff_by_campaign(cfg)
        user_dates     = _load_sync_dates_by_username(cfg, key_list)
    else:
        if not cfg["project_type_id"]:
            log.warning("[cdd_sync] is_admin_console=FALSE but project_type_id is empty — skipping")
            return None
        cdds, key_list = _load_staff_by_project_type(cfg)
        user_dates     = _load_sync_dates_by_user_id(cfg, key_list)

    if not cdds:
        log.warning("[cdd_sync] 0 CDDs found — check campaign_number / project_type_id")
        return None

    all_rows, day_labels = _build_cdd_rows(cdds, user_dates, cfg, is_admin)

    # Cutoff stats are a TODAY-only operational metric (same-day UTC cutoff), so
    # they are meaningless in a whole-campaign cumulative run and skipped there.
    if cfg.get("cumulative"):
        synced_by_1700 = synced_by_1730 = synced_keys_1730 = None
    else:
        synced_by_1700   = _count_synced_by_cutoff(cfg, 17, 0)
        synced_by_1730   = _count_synced_by_cutoff(cfg, 17, 30)
        synced_keys_1730 = _get_synced_keys_by_cutoff(cfg, 17, 30)

    log.info(f"[cdd_sync:collect] {len(all_rows)} CDDs, {len(day_labels)} campaign days")
    return {
        "is_admin":         is_admin,
        "all_rows":         all_rows,
        "day_labels":       day_labels,
        "synced_by_1700":   synced_by_1700,
        "synced_by_1730":   synced_by_1730,
        "synced_keys_1730": sorted(synced_keys_1730) if synced_keys_1730 is not None else None,
    }


def render(cfg, data):
    """Stage 2 (pure): build the CDD sync workbook from collected data."""
    all_rows   = data["all_rows"]
    day_labels = data["day_labels"]
    is_admin   = data["is_admin"]

    COLS = (["LGA", "Health Facility", "Username", "User ID", "Days Synced"]
            + list(day_labels.values()) + ["Status", "Sync Dates"])

    df_all = pd.DataFrame(all_rows).sort_values(["LGA", "Days Synced"])
    for col in COLS:
        if col not in df_all.columns:
            df_all[col] = ""

    wb = Workbook()
    wb.remove(wb.active)

    ws_sum = wb.create_sheet("SUMMARY")
    _write_summary_tab(ws_sum, all_rows, cfg)

    total_cdds_count = len(all_rows)
    last_row = ws_sum.max_row + 2
    time_rows = []
    if data["synced_by_1700"] is not None:
        pct = f"{data['synced_by_1700']/total_cdds_count*100:.1f}%" if total_cdds_count else "-"
        time_rows.append(("Synced by 17:00 today (UTC)", data["synced_by_1700"], pct))
    if data["synced_by_1730"] is not None:
        pct = f"{data['synced_by_1730']/total_cdds_count*100:.1f}%" if total_cdds_count else "-"
        time_rows.append(("Synced by 17:30 today (UTC)", data["synced_by_1730"], pct))
    for i, (label, count, pct) in enumerate(time_rows):
        r = last_row + i
        ws_sum.cell(r, 1, label)
        ws_sum.cell(r, 2, count)
        ws_sum.cell(r, 3, pct)
        for ci in range(1, 4):
            style_sync_cell(ws_sum.cell(r, ci), fill=SYNC_TOTAL_FILL, bold=True)

    for lga in sorted(df_all["LGA"].unique()):
        df_lga = df_all[df_all["LGA"] == lga][COLS].copy().reset_index(drop=True)
        df_lga.insert(0, "#", range(1, len(df_lga) + 1))
        safe = str(lga).translate(str.maketrans("/\\*?[]:", "-------"))[:31] or "Unknown"
        _write_dataframe_tab(wb.create_sheet(safe), df_lga)

    df_never = df_all[df_all["Status"] == "NEVER SYNCED"][
        ["LGA", "Health Facility", "Username", "User ID"]
    ].reset_index(drop=True)
    df_never.insert(0, "#", range(1, len(df_never) + 1))
    _write_dataframe_tab(wb.create_sheet("NEVER SYNCED"), df_never)

    df_low = df_all[df_all["Status"] == "LOW"][
        ["LGA", "Health Facility", "Username", "User ID", "Days Synced", "Sync Dates", "Status"]
    ].reset_index(drop=True)
    df_low.insert(0, "#", range(1, len(df_low) + 1))
    _write_dataframe_tab(wb.create_sheet("LOW SYNCED"), df_low)

    synced_keys_1730 = data["synced_keys_1730"]
    if synced_keys_1730 is not None:
        synced_keys = set(synced_keys_1730)
        lookup_key  = "Username" if is_admin else "User ID"
        not_synced  = [r for r in all_rows if r[lookup_key].lower() not in synced_keys]
        df_ns = pd.DataFrame(not_synced if not_synced else [],
                             columns=["LGA", "Health Facility", "Username", "User ID",
                                      "Days Synced", "Sync Dates", "Status"])[
            ["LGA", "Health Facility", "Username", "User ID"]
        ].reset_index(drop=True)
        df_ns.insert(0, "#", range(1, len(df_ns) + 1))
        _write_dataframe_tab(wb.create_sheet("NOT SYNCED BY 1730"), df_ns)
        log.info(f"  not synced by 17:30 UTC: {len(not_synced):,} CDDs")

    out = cfg["sync_xlsx"]
    wb.save(out)

    never_count = sum(1 for r in all_rows if r["Status"] == "NEVER SYNCED")
    log.info(f"[cdd_sync:render] saved -> {out}  ({len(all_rows)} CDDs, {never_count} never synced)")
    return out


def run(cfg):
    log.info(f"[cdd_sync] {cfg['state_name']} Day {cfg['DAY']} ...")
    data = collect(cfg)
    if data is None:
        return None
    save_checkpoint(cfg, "cdd_sync", data)
    return render(cfg, data)


def rerun_from_checkpoint(cfg):
    """Rebuild the CDD sync Excel from the saved checkpoint — no ES access needed."""
    return render(cfg, load_checkpoint(cfg, "cdd_sync"))
