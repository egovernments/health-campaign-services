"""
excel_style.py — shared, best-effort styling for generated report sheets.

Call style_report_workbook(path) AFTER a report writes its .xlsx (works for both
pandas to_excel and raw openpyxl output). It styles the header row (bold, coloured
fill, white font), auto-sizes columns to fit content (capped), freezes the header,
and adds an auto-filter — applied to every worksheet.

Best-effort by design: any failure is logged and swallowed so styling can never
block report generation. Column widths are sampled (header + first N rows) to stay
O(cols) rather than scanning every cell of a large report.
"""
import os

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL_COLOR = "93C47D"   # DIGIT light green (matches excel-ingestion default header)
HEADER_FONT_COLOR = "000000"   # black — readable on the light-green fill
HEADER_FONT_SIZE = 12
HEADER_ROW_HEIGHT = 28
MIN_WIDTH = 12
MAX_WIDTH = 60
PAD = 2
WIDTH_SAMPLE_ROWS = 200        # rows scanned for width sizing (keeps big reports fast)
WRAP_MAX_ROWS = 100000         # skip per-cell wrap beyond this many rows (perf guard)


def _style_sheet(ws):
    if ws.max_row < 1 or ws.max_column < 1:
        return

    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=HEADER_FONT_SIZE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header row = row 1 (both to_excel(index=False) and sheet.append(header) put it there).
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # Column widths: longest of header + a sample of data rows, clamped. Columns whose
    # content exceeds MAX_WIDTH are "wide" - their data cells get wrapped below so long
    # values stay readable instead of being clipped at the cap.
    sample_limit = min(ws.max_row, WIDTH_SAMPLE_ROWS + 1)
    wide_cols = []
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        longest = 0
        for row in range(1, sample_limit + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                longest = max(longest, len(str(v)))
        ws.column_dimensions[letter].width = max(MIN_WIDTH, min(longest + PAD, MAX_WIDTH))
        if longest + PAD > MAX_WIDTH:
            wide_cols.append(col)

    # Wrap data cells of wide columns (Excel auto-grows row height for wrapped cells).
    # Bounded to wide columns + capped by row count so a huge report isn't slowed down.
    if wide_cols and 1 < ws.max_row <= WRAP_MAX_ROWS:
        wrap_align = Alignment(wrap_text=True, vertical="top")
        for row in range(2, ws.max_row + 1):
            for col in wide_cols:
                ws.cell(row=row, column=col).alignment = wrap_align

    # Freeze the header and enable filtering.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Protect the sheet: cells become read-only (can't be edited), but keep it
    # usable for viewers - selecting, sorting, filtering and column/row resizing
    # stay allowed. Optional password (REPORT_SHEET_PASSWORD) makes unprotecting
    # deliberate; without it the sheet is still read-only by default in Excel.
    ws.protection.sheet = True
    password = os.getenv("REPORT_SHEET_PASSWORD")
    if password:
        ws.protection.password = password
    # Allow every view/format convenience (True=blocked, False=allowed in OOXML). Only
    # editing cell *content* (locked cells) and inserting/deleting rows/columns stay
    # blocked, so the data can't be tampered with but the sheet is fully usable. These
    # take effect with or without a password (password only gates unprotecting).
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False


def style_report_workbook(path):
    """Style every worksheet of the workbook at `path` in place. Never raises."""
    try:
        wb = load_workbook(path)
        for ws in wb.worksheets:
            _style_sheet(ws)
        wb.save(path)
        print(f"[EXCEL_STYLE] styled {path}")
    except Exception as e:
        print(f"[EXCEL_STYLE] skipped styling for {path}: {e}")
