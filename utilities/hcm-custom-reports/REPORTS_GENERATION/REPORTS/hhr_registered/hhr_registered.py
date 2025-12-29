import re
import warnings
import requests
import json
import time
import copy
import datetime
from openpyxl import Workbook
import os
import sys
import argparse

# === COMMAND LINE ARGUMENTS ===
parser = argparse.ArgumentParser()
parser.add_argument('--campaign_identifier', required=True,
                    help='Campaign identifier (can be campaignNumber or projectTypeId)')
parser.add_argument('--identifier_type', default='campaignNumber',
                    help='Type of identifier: "campaignNumber" or "projectTypeId"')
parser.add_argument('--start_date', default='')
parser.add_argument('--end_date', default='')
parser.add_argument('--file_name', required=True)
args = parser.parse_args()

CAMPAIGN_IDENTIFIER = args.campaign_identifier
IDENTIFIER_TYPE = args.identifier_type
START_DATE = args.start_date
END_DATE = args.end_date
FILE_NAME = args.file_name

# === PATH SETUP ===
file_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.append(file_path)
from REPORTS_GENERATION.COMMON_UTILS.custom_date_utils import get_custom_dates_of_reports

ES_HOUSEHOLD_SEARCH="http://elasticsearch-master.es-upgrade.svc.cluster.local:9200/household-index-v1/_search"
ES_HOUSEHOLD_MEMBER_SEARCH="http://elasticsearch-master.es-upgrade.svc.cluster.local:9200/household-member-index-v1/_search"
ES_INDIVIDUAL_SEARCH = "http://elasticsearch-master.es-upgrade.svc.cluster.local:9200/individual-index-v1/_search"
ES_BENEFICIARY_SEARCH="http://elasticsearch-master.es-upgrade.svc.cluster.local:9200/project-beneficiary-index-v1/_search"

max_retries = 60
retry_delay = 5

# Filter out the specific warning messages by message
warnings.filterwarnings("ignore", message="Unverified HTTPS request is being made.*")
users_file = open("USERS/users_burundi.json")
users_json = json.load(users_file)

boundary_file = open("BOUNDARY/BOUNDARY_bi.json")
boundary_json = json.load(boundary_file)

localization_data = open("BOUNDARY/localization_data.json")
localization_data_json = json.load(localization_data)

def get_resp(url, data, es=False):
    failed = False
    for _ in range(max_retries):
        if failed:
            print(_, f" retry count where max retry count is {max_retries}")
        try:
            if es:
                response = requests.post(url, data=json.dumps(data), headers={"Content-Type": "application/json", "Authorization": "Basic ZWxhc3RpYzpaRFJsT0RJME1UQTNNV1ppTVRGbFptRms="}, verify=False)
            else:
                response = requests.post(url, data=json.dumps(data), headers={"Content-Type": "application/json"})
            
            if response.status_code == 200:
                print(response)
                return response
            print(response.json())
        except requests.exceptions.ConnectionError:
            print(f"Connection error. Retrying in {retry_delay} seconds...")
            print(f"Retrying connecting to {url}")
            failed = True
        time.sleep(retry_delay)



def load_user_info(roles, merge=False):
    return_users_by_role = []
    return_users = {}
    if type(roles) is str:
        with open( "USERS/" + roles + "S.json", "r") as role_file:
            users = json.load(role_file)
        return users
    for role in roles:
        with open("USERS/" + role + "S.json", "r") as role_file:
            users = json.load(role_file)
            if not merge:
                return_users_by_role.append(users)
            else:
                return_users = {**return_users, **users}
    return return_users if merge else return_users_by_role


def get_user_details(user_id):
    users_info = load_user_info(["USER"], merge=True)
    mobile_number = users_info.get(user_id, {}).get("number", "")
    return mobile_number


# === DATE RANGE ===
lteTime, gteTime, start_date_str, end_date_str = get_custom_dates_of_reports(START_DATE, END_DATE)

# === CAMPAIGN FILTER FIELD ===
if IDENTIFIER_TYPE == "projectTypeId":
    CAMPAIGN_FILTER_FIELD = "Data.projectTypeId.keyword"
    print(f"Using projectTypeId filter: {CAMPAIGN_IDENTIFIER}")
else:
    CAMPAIGN_FILTER_FIELD = "Data.campaignNumber.keyword"
    print(f"Using campaignNumber filter: {CAMPAIGN_IDENTIFIER}")


def extract_data():
    query = {
        "size": 10000,
            "query": {
                "bool": {
                    "must": [
                        {
                            "term": {
                                CAMPAIGN_FILTER_FIELD: CAMPAIGN_IDENTIFIER
                            }
                        },
                        {
                            "range": {
                                "Data.@timestamp": {
                                "gte": gteTime,
                                "lte": lteTime
                                }
                            }
                        }
                    ]
                }
            },
            "sort" : [
                {
                    "Data.@timestamp" : "asc"
                },
                {
                    "Data.household.clientAuditDetails.createdTime" : "asc"
                }
            ],
        "_source": ["Data.boundaryHierarchy", "Data.household.clientReferenceId", "Data.userName", "Data.household.clientAuditDetails.createdBy", "Data.nameOfUser", "Data.household.clientAuditDetails.createdTime", "Data.syncedTimeStamp", "Data.household.memberCount","Data.household.address.locality.code"]
    }
    docs = get_resp(ES_HOUSEHOLD_SEARCH, query, True).json()["hits"]["hits"]
    print(f"{len(docs)}, docs from household index")
    # print(docs)
    household_registered = {}
    clientReferenceIds = []
    while len(docs) > 0:
        last_sort_value = docs[-1]["sort"][0]
        last_createdTime = docs[-1]["sort"][1]
        print(f"Last sort value: {last_sort_value}")

        for item in docs:
            locCode = item["_source"]["Data"]["household"]["address"]["locality"]["code"]
            ddp = get_distribution_point(locCode)
            household_registered[item["_source"]["Data"]["household"]["clientReferenceId"]] = {
                "boundaryHierarchy" : item["_source"]["Data"]["boundaryHierarchy"] ,
                "nameOfUser" : item["_source"]["Data"]["nameOfUser"],
                "userName" : item["_source"]["Data"]["userName"],
                "userId" : item["_source"]["Data"]["household"]["clientAuditDetails"]["createdBy"],
                "syncedTime" : item["_source"]["Data"]["syncedTimeStamp"],
                "createdTime" : item["_source"]["Data"]["household"]["clientAuditDetails"]["createdTime"],
                "memberCount" : item["_source"]["Data"]["household"]["memberCount"],
                "ddp": ddp,
                "localityCode" : item["_source"]["Data"]["household"]["address"]["locality"]["code"]
            }
            clientReferenceIds.append(item["_source"]["Data"]["household"]["clientReferenceId"])
        print(f"{len(household_registered)} objects in household registered")

        query_with_search_after = copy.deepcopy(query)
        query_with_search_after["search_after"] = [last_sort_value, last_createdTime]
        response = get_resp(ES_HOUSEHOLD_SEARCH, query_with_search_after, True).json()
        docs = response["hits"]["hits"]
        print(f"{len(docs)} docs from household index")
    print(f"{len(household_registered)} objects in household registered")
    print(f"{len(clientReferenceIds)} number of clientreferenceIds")
    return household_registered, clientReferenceIds

def convert_epoch_to_datetime(epoch_time):
    epoch_seconds = epoch_time / 1000

    date_time = datetime.datetime.fromtimestamp(epoch_seconds)

    # Format datetime object as "dd-mm-yy hh:mm:ss"
    formatted_date_time = date_time.strftime('%d-%m-%y %H:%M:%S')

    return formatted_date_time

def format_date(date_string):
    date_format = "%Y-%m-%dT%H:%M:%S.%fZ"

    date_time = datetime.datetime.strptime(date_string, date_format)

    # Format datetime object as "dd-mm-yy hh:mm:ss"
    formatted_date_time = date_time.strftime('%d-%m-%y %H:%M:%S')

    return formatted_date_time

def get_distribution_point(locality_code):
    if locality_code is not None and locality_code != "null":
        for obj in localization_data_json["messages"]:
            if obj["code"] == locality_code+"_DP":
                return obj["message"]
        
    return "null"

def get_household_member(clientReferenceIds) :
    print(f"{len(clientReferenceIds)} number of clientReferenceIds provided to fetch household members")
    data = {}
    individualClientReferenceIds = []

    # Split clientReferenceIds into chunks of 10000 or less
    chunk_size = 10000
    for i in range(0, len(clientReferenceIds), chunk_size):
        chunk = clientReferenceIds[i:i + chunk_size]
        query = {
            "size" : len(chunk),
            "query": {
                "bool": {
                    "must": [
                                {
                                "terms": {
                                    "Data.householdMember.householdClientReferenceId.keyword": chunk
                            }
                        },
                        {
                            "match": {
                                "Data.householdMember.isHeadOfHousehold": True
                            }
                        }
                    ]
                }
            },
            "_source": ["Data.householdMember.individualClientReferenceId", "Data.householdMember.householdClientReferenceId"]
        }
        docs = get_resp(ES_HOUSEHOLD_MEMBER_SEARCH, query, True).json()["hits"]["hits"]
        for item in docs:
            data[item["_source"]["Data"]["householdMember"]["householdClientReferenceId"]] = {
                "individualClientReferenceId" : item["_source"]["Data"]["householdMember"]["individualClientReferenceId"]
            }
            individualClientReferenceIds.append(item["_source"]["Data"]["householdMember"]["individualClientReferenceId"])
    print(f"{len(data)} number of modified objects from household member index")
    return individualClientReferenceIds, data


def get_individuals(individualClientReferenceIds):
    print(f"{len(individualClientReferenceIds)} number of individualClientReferenceIds provided to fetch individuals")
    data = {}

    # Split individualClientReferenceIds into chunks of 10000 or less
    chunk_size = 10000
    for i in range(0, len(individualClientReferenceIds), chunk_size):
        chunk = individualClientReferenceIds[i:i + chunk_size]
        query = {
            "size": len(chunk),
            "query": {
                "terms": {
                    "_id": chunk
                }
            },
            "_source": ["gender", "name", "clientAuditDetails.createdBy", "mobileNumber", "address"]
        }
        docs = get_resp(ES_INDIVIDUAL_SEARCH, query, True).json()["hits"]["hits"]
        for item in docs:
            data[item["_id"]] = {
                "gender" : item["_source"]["gender"],
                "firstName" : item["_source"]["name"]["givenName"],
                "lastName" :  item["_source"]["name"]["familyName"],
                "latitude" : item["_source"]["address"][0]["latitude"],
                "longitude" : item["_source"]["address"][0]["longitude"]
                # "individualMobileNumber" : get_decrypted_mobile_number(item["_source"]["mobileNumber"]),
            }            
    print(f"{len(data)} number of modified objects from individual index")
    return data

def get_voucher_codes(clientReferenceIds):
    print(f"{len(clientReferenceIds)} number of clientReferenceIds provided to fetch voucher codes")
    data = {}

    chunk_size = 10000
    for i in range(0, len(clientReferenceIds), chunk_size):
        chunk = clientReferenceIds[i:i + chunk_size]
        query = {
            "size": len(chunk),
            "query": {
                "terms": {
                    "beneficiaryClientReferenceId.keyword": chunk
                }
            },
            "_source": ["tag", "beneficiaryClientReferenceId"]
        }
        docs = get_resp(ES_BENEFICIARY_SEARCH, query, True).json()["hits"]["hits"]
        for item in docs:
            data[item["_source"]["beneficiaryClientReferenceId"]] = {
                "voucherCode" : item["_source"]["tag"]
            }

    print(f"{len(data)} number of modified objects from project beneficiary index")
    return data


def get_individual_client_referenceids():
    mapping, household_clientreferenceIds = extract_data()
    individualClientReferenceIds, household_memeber_info = get_household_member(household_clientreferenceIds)
    voucher_codes = get_voucher_codes(household_clientreferenceIds)

    for client_reference_id, value in mapping.items():
        if client_reference_id in household_memeber_info:
            mapping[client_reference_id]["individualClientReferenceId"] = household_memeber_info[client_reference_id]["individualClientReferenceId"]

    for client_reference_id, value in mapping.items():
        if client_reference_id in voucher_codes:
            mapping[client_reference_id]["voucherCode"] = voucher_codes[client_reference_id]["voucherCode"]

    count = 0
    for client_reference_id, value in mapping.items():
        if mapping[client_reference_id].get("individualClientReferenceId") is None:
            count += 1
    print("no of hh which have no hhm(individual hh head names) mapped - ")
    print(count)

    return individualClientReferenceIds, mapping

def get_individual_info():
    individualClientReferenceIds, mapping = get_individual_client_referenceids()
    individual_info = get_individuals(individualClientReferenceIds)

    for client_ref_id, values in mapping.items():
        if "individualClientReferenceId" in values:
            individualClientReferenceId = values["individualClientReferenceId"]
            if individualClientReferenceId in individual_info:
                values["gender"] = individual_info[individualClientReferenceId]["gender"]
                values["firstName"] = individual_info[individualClientReferenceId]["firstName"]
                values["lastName"] = individual_info[individualClientReferenceId]["lastName"]
                values["latitude"] = individual_info[individualClientReferenceId]["latitude"]
                values["longitude"] = individual_info[individualClientReferenceId]["longitude"]
    return mapping

def generate_report():
    data = get_individual_info()
    
    # Save raw data to JSON for reference
    # with open("REPORTS_GENERATION/REPORTS/hhr_registered/data.json", "w") as data_json:
    #     json.dump(data, data_json, indent=4)

    workbook = Workbook()
    sheets = {}
    # sheet = workbook.active

    header_row = [
        "Province", "District", "Administrative Province", "Locality", "Village", "Designated Distribution Point",
        "HHR Username", "HHR Phone number", "HHR Name","Household Head Name", "Gender",
        "Number of People Living in HHs", "Serial Number of Voucher", "Latitude", "Longitude",
        "Created Time", "Synced Time"
    ]
    # sheet.append(header_row)

    for id, value in data.items():
        mobile_number = "null"
        if len(value.get("userId", "")) > 0:
            # mobile_number = get_user_details(value["userId"])
            mobile_number = users_json.get(value["userId"],'')

        # if all(field in value and value[field] is not None for field in ["voucherCode", "firstName", "lastName"]):
            
        boundary = value.get("boundaryHierarchy", {})
        province = boundary.get("province", "null")
        district = boundary.get("district", "null")
        administrativeProvince = boundary.get("administrativeProvince", "null")
        locality = boundary.get("locality", "null")
        village = boundary.get("village", "null")
        created_time = convert_epoch_to_datetime(value["createdTime"])
        synced_time = value["syncedTime"]
        # distribution_point = get_distribution_point(value["localityCode"])
        distribution_point = value["ddp"] #Todo
        row_data = [
                province, district, administrativeProvince, locality, village, distribution_point,
                value["userName"], mobile_number, value["nameOfUser"],
                f"{value['firstName']} {value['lastName']}" if "firstName" in value and "lastName" in value else "null",
                value.get("gender", "null"), value.get("memberCount", "null"), value.get("voucherCode", "null"),
                value.get("latitude", "null"), value.get("longitude", "null"), created_time, synced_time
            ]
        print("row_data",row_data)
        if locality not in sheets:
            sheet = workbook.create_sheet(title=locality[:31])  # Excel sheet names are limited to 31 chars
            sheet.append(header_row)
            sheets[locality] = sheet

        sheets[locality].append(row_data)
        # sheet.append(row_data)
    # Remove default sheet if unused
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        workbook.remove(workbook["Sheet"])
    # Save workbook - use FILE_NAME from command line args
    output_file = f"{FILE_NAME}.xlsx"
    output_path = os.path.join(os.getcwd(), output_file)
    workbook.save(output_path)
    print(f"Report generated: {output_path}")

generate_report()