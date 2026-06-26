import os
import sys
import warnings
import argparse
from concurrent.futures import ThreadPoolExecutor
from collections import defaultdict

# === COMMAND LINE ARGUMENTS ===
parser = argparse.ArgumentParser()
parser.add_argument('--campaign_identifier', required=True,
                    help='Campaign identifier (can be campaignNumber or projectTypeId)')
parser.add_argument('--identifier_type', default='campaignNumber',
                    help='Type of identifier: "campaignNumber" or "projectTypeId"')
parser.add_argument('--start_date', default='')
parser.add_argument('--end_date', default='')
parser.add_argument('--file_name', required=True)
args = parser.parse_args()

CAMPAIGN_IDENTIFIER = args.campaign_identifier
IDENTIFIER_TYPE     = args.identifier_type
START_DATE          = args.start_date
END_DATE            = args.end_date
FILE_NAME           = args.file_name

# === PATH SETUP ===
file_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.append(file_path)

warnings.filterwarnings("ignore", message="Unverified HTTPS request is being made.*")

from COMMON_UTILS.custom_date_utils import get_custom_dates_of_reports
from COMMON_UTILS.common_utils import get_resp, es_index_url, es_scroll_url

# === CONSTANTS ===
ES_PROJECT_TASK_INDEX = es_index_url("project-task-index-v1")
ES_INDIVIDUAL_INDEX   = es_index_url("individual-index-v1")
ES_HHM_V1             = es_index_url("household-member-index-v1")
ES_SCROLL_API         = es_scroll_url()

SCROLL_SIZE       = 10000
BATCH_SIZE        = 5000
MAX_FETCH_WORKERS = 8

RELEVANT_STATUSES = [
    "ADMINISTRATION_SUCCESS",
    "VISITED",
    "INELIGIBLE",
    "BENEFICIARY_MIGRATED",
    "BENEFICIARY_DIED",
    "BENEFICIARY_ABSENT",
    "BENEFICIARY_REFUSED",
]

# Maps each status string to a small int (0-6).
# Used to build compact int keys for duplicate detection — avoids storing
# per-individual nested dicts which would be ~1.8 GB for 2.6M individuals.
_STATUS_IDX = {s: i for i, s in enumerate(RELEVANT_STATUSES)}

EXPORT_COLUMNS = [
    "Individual ID", "Country", "State", "LGA", "Ward", "Health Facility",
    "Username", "Child Name", "Age", "Gender",
    "Child Beneficiary ID", "Household Head Name",
    "Latitude", "Longitude", "Product Name", "Date of Administration",
    "Cycle Index", "Administration Status", "Timestamp", "Record Type",
    "Quantity Administered", "Redose Quantity Administered",
]

# === DATE RANGE ===
lteTime, gteTime, start_date_str, end_date_str = get_custom_dates_of_reports(START_DATE, END_DATE)

# === CAMPAIGN FILTER ===
if IDENTIFIER_TYPE == "projectTypeId":
    CAMPAIGN_FILTER_FIELD = "Data.projectTypeId.keyword"
    print(f"Using projectTypeId filter: {CAMPAIGN_IDENTIFIER}")
else:
    CAMPAIGN_FILTER_FIELD = "Data.campaignNumber.keyword"
    print(f"Using campaignNumber filter: {CAMPAIGN_IDENTIFIER}")

print(f"Reports start date : {start_date_str}")
print(f"Reports end date   : {end_date_str}")
print(f"Campaign           : {CAMPAIGN_IDENTIFIER}")
print(f"\n===== Generating report : CHILDREN_TREATED_REPORT_UPDATED\n")


# === PARALLEL FETCH HELPER ===
def parallel_fetch(batches, fetch_fn):
    out = []
    with ThreadPoolExecutor(max_workers=MAX_FETCH_WORKERS) as ex:
        for res in ex.map(fetch_fn, batches):
            out.append(res)
    return out


# ====================================================================
# PASS 1 — Minimal scroll: only individualId + administrationStatus.
#
# Duplicate detection uses two int sets instead of a nested defaultdict.
# For 2.6M unique individuals a nested defaultdict costs ~1.8 GB;
# two int sets cost ~200 MB peak (seen_keys is freed right after).
#
# Key: hash((interned_uuid, status_idx))
#   status_idx 0-6 → CPython caches those ints (no heap allocation).
#   Tuple hash is well-distributed; collision probability < 10^-7.
#   Hash randomisation is per-process so keys stay consistent within
#   a single run (Pass 1 and Pass 2 share the same Python process).
# ====================================================================
def pass1_count_records():
    query = {
        "size": SCROLL_SIZE,
        "query": {
            "bool": {
                "must": [
                    {"term":  {CAMPAIGN_FILTER_FIELD: CAMPAIGN_IDENTIFIER}},
                    {"range": {"Data.@timestamp": {"gte": gteTime, "lte": lteTime}}},
                    {"terms": {"Data.administrationStatus.keyword": RELEVANT_STATUSES}},
                ]
            }
        },
        "_source": ["Data.individualId", "Data.administrationStatus"],
    }

    scroll_url = f"{ES_PROJECT_TASK_INDEX}?scroll=10m"
    scroll_id  = None
    total      = 0

    # seen_keys  — every (ind_id, status_idx) pair seen at least once
    # dup_keys   — pairs seen more than once → these records are "duplicate"
    # ind_ids_set — unique interned individual UUID strings (for enrichment)
    # seen_keys is deleted after Pass 1; dup_keys + ind_ids_set are kept.
    seen_keys   = set()
    dup_keys    = set()
    ind_ids_set = set()

    print("[Pass 1/2] Counting records per individual...")

    while True:
        if scroll_id is None:
            resp = get_resp(scroll_url, query, True).json()
        else:
            resp = get_resp(ES_SCROLL_API, {"scroll": "10m", "scroll_id": scroll_id}, True).json()

        scroll_id = resp.get("_scroll_id", "")
        hits = resp.get("hits", {}).get("hits", [])
        if not hits:
            break

        for raw in hits:
            doc    = raw["_source"]["Data"]
            ind_id = doc.get("individualId")
            if not ind_id:
                continue

            status_idx = _STATUS_IDX.get(doc.get("administrationStatus", ""))
            if status_idx is None:
                continue

            # sys.intern deduplicates the UUID string object so the same
            # string is shared across all dicts built in this process.
            iid = sys.intern(ind_id)
            ind_ids_set.add(iid)
            k = hash((iid, status_idx))
            if k in seen_keys:
                dup_keys.add(k)
            else:
                seen_keys.add(k)

        total += len(hits)
        if total % 100_000 == 0:
            print(f"  {total:,} counted...")

    n_seen = len(seen_keys)
    del seen_keys  # free ~200 MB before enrichment phase

    print(f"  Pass 1 done: {total:,} records | {n_seen:,} unique (ind, status) pairs | {len(ind_ids_set):,} unique individuals\n")
    return dup_keys, ind_ids_set, total


# ====================================================================
# ENRICHMENT — Build flat lookup dicts from individual + HHM indices.
# Returns:
#   ind_id_to_name      : {clientReferenceId → full name}
#   ind_id_to_head_name : {clientReferenceId → household head full name}
# ====================================================================
def build_enrichment_maps(ind_ids_list):
    # ── A: individual → name ────────────────────────────────────────
    print(f"[Enrich 1/2] Fetching child names ({len(ind_ids_list):,} individuals)...")

    def fetch_individual_batch(batch):
        q = {
            "size": BATCH_SIZE,
            "query": {"terms": {"clientReferenceId.keyword": batch}},
            "_source": ["name", "clientReferenceId"],
        }
        return get_resp(ES_INDIVIDUAL_INDEX, q, True).json().get("hits", {}).get("hits", [])

    ind_id_to_name = {}
    batches = [ind_ids_list[i:i + BATCH_SIZE] for i in range(0, len(ind_ids_list), BATCH_SIZE)]
    for results in parallel_fetch(batches, fetch_individual_batch):
        for doc in results:
            src  = doc["_source"]
            n    = src.get("name") or {}
            name = f"{n.get('givenName') or ''} {n.get('familyName') or ''}".strip()
            # sys.intern: UUID key is shared with the same object interned in Pass 1
            ind_id_to_name[sys.intern(src["clientReferenceId"])] = name

    print(f"  Matched {len(ind_id_to_name):,} / {len(ind_ids_list):,} individuals.")

    # ── B: individual → household id ────────────────────────────────
    print(f"[Enrich 2/2] Fetching household head names...")

    def fetch_hh_member_batch(batch):
        q = {
            "size": BATCH_SIZE,
            "query": {"terms": {"Data.householdMember.individualClientReferenceId.keyword": batch}},
            "_source": [
                "Data.householdMember.individualClientReferenceId",
                "Data.householdMember.householdClientReferenceId",
            ],
        }
        return get_resp(ES_HHM_V1, q, True).json().get("hits", {}).get("hits", [])

    ind_to_hh = {}
    for results in parallel_fetch(batches, fetch_hh_member_batch):
        for doc in results:
            hh = doc["_source"]["Data"]["householdMember"]
            # intern both sides so keys are shared with ind_id_to_name
            ind_to_hh[sys.intern(hh["individualClientReferenceId"])] = sys.intern(hh["householdClientReferenceId"])

    # ── C: household id → head individual id ────────────────────────
    hh_ids = list(set(ind_to_hh.values()))
    print(f"  Unique households: {len(hh_ids):,}")

    def fetch_hh_head_batch(batch):
        q = {
            "size": BATCH_SIZE,
            "query": {
                "bool": {
                    "must": [
                        {"terms": {"Data.householdMember.householdClientReferenceId.keyword": batch}},
                        {"term":  {"Data.householdMember.isHeadOfHousehold": True}},
                    ]
                }
            },
            "_source": [
                "Data.householdMember.householdClientReferenceId",
                "Data.householdMember.individualClientReferenceId",
            ],
        }
        return get_resp(ES_HHM_V1, q, True).json().get("hits", {}).get("hits", [])

    hh_to_head_ref = {}
    hh_batches = [hh_ids[i:i + BATCH_SIZE] for i in range(0, len(hh_ids), BATCH_SIZE)]
    for results in parallel_fetch(hh_batches, fetch_hh_head_batch):
        for doc in results:
            hh = doc["_source"]["Data"]["householdMember"]
            hh_to_head_ref[hh["householdClientReferenceId"]] = hh["individualClientReferenceId"]

    # ── D: fetch names for heads not already in ind_id_to_name ──────
    unknown_heads = [r for r in set(hh_to_head_ref.values()) if r not in ind_id_to_name]
    if unknown_heads:
        def fetch_head_name_batch(batch):
            q = {
                "size": BATCH_SIZE,
                "query": {"terms": {"clientReferenceId.keyword": batch}},
                "_source": ["clientReferenceId", "name"],
            }
            return get_resp(ES_INDIVIDUAL_INDEX, q, True).json().get("hits", {}).get("hits", [])

        head_batches = [unknown_heads[i:i + BATCH_SIZE] for i in range(0, len(unknown_heads), BATCH_SIZE)]
        for results in parallel_fetch(head_batches, fetch_head_name_batch):
            for doc in results:
                src  = doc["_source"]
                n    = src.get("name") or {}
                name = f"{n.get('givenName') or ''} {n.get('familyName') or ''}".strip()
                ind_id_to_name[sys.intern(src["clientReferenceId"])] = name

    # ── E: build final ind_id → head_name flat dict ─────────────────
    ind_id_to_head_name = {}
    for ind_id, hh_id in ind_to_hh.items():
        head_ref = hh_to_head_ref.get(hh_id, "")
        if head_ref:
            ind_id_to_head_name[ind_id] = ind_id_to_name.get(head_ref, "")

    # Free large intermediate dicts before Pass 2
    del ind_to_hh, hh_to_head_ref

    print(f"  Head names resolved: {len(ind_id_to_head_name):,}\n")
    return ind_id_to_name, ind_id_to_head_name


# ====================================================================
# PASS 2 — Full scroll, stream rows directly to Excel.
# openpyxl write_only keeps memory constant regardless of row count.
# Tracks quality/status stats inline — no DataFrame needed.
# ====================================================================
def pass2_stream_to_excel(dup_keys, ind_id_to_name, ind_id_to_head_name, output_path):
    from openpyxl import Workbook

    query = {
        "size": SCROLL_SIZE,
        "query": {
            "bool": {
                "must": [
                    {"term":  {CAMPAIGN_FILTER_FIELD: CAMPAIGN_IDENTIFIER}},
                    {"range": {"Data.@timestamp": {"gte": gteTime, "lte": lteTime}}},
                    {"terms": {"Data.administrationStatus.keyword": RELEVANT_STATUSES}},
                ]
            }
        },
        # Sort by cycleIndex so the Excel output is grouped by cycle
        "sort": [
            {"Data.additionalDetails.cycleIndex.keyword": {"order": "asc", "missing": "_last"}},
            {"Data.@timestamp": {"order": "asc"}},
        ],
        "_source": [
            "Data.boundaryHierarchy",
            "Data.age", "Data.individualId",
            "Data.@timestamp", "Data.taskDates", "Data.userName",
            "Data.quantity", "Data.productName", "Data.administrationStatus",
            "Data.latitude", "Data.longitude", "Data.additionalDetails",
        ],
    }

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append(EXPORT_COLUMNS)

    scroll_url  = f"{ES_PROJECT_TASK_INDEX}?scroll=10m"
    scroll_id   = None
    total       = 0
    status_counts      = defaultdict(int)
    record_type_counts = defaultdict(lambda: {"unique": 0, "duplicate": 0})
    missing            = {"Child Name": 0, "Child Beneficiary ID": 0, "Household Head Name": 0}

    print("[Pass 2/2] Streaming records to Excel...")

    while True:
        if scroll_id is None:
            resp = get_resp(scroll_url, query, True).json()
        else:
            resp = get_resp(ES_SCROLL_API, {"scroll": "10m", "scroll_id": scroll_id}, True).json()

        scroll_id = resp.get("_scroll_id", "")
        hits = resp.get("hits", {}).get("hits", [])
        if not hits:
            break

        for raw in hits:
            doc        = raw["_source"]["Data"]
            ind_id     = doc.get("individualId")
            if not ind_id:
                continue

            admin_status   = doc.get("administrationStatus", "")
            additional     = doc.get("additionalDetails") or {}
            boundary       = doc.get("boundaryHierarchy") or {}
            quantity       = doc.get("quantity")

            child_name     = ind_id_to_name.get(ind_id, "")
            head_name      = ind_id_to_head_name.get(ind_id, "")
            beneficiary_id = additional.get("beneficiaryId", "")
            status_idx  = _STATUS_IDX.get(admin_status)
            record_type = "duplicate" if status_idx is not None and hash((ind_id, status_idx)) in dup_keys else "unique"

            if not child_name:     missing["Child Name"] += 1
            if not beneficiary_id: missing["Child Beneficiary ID"] += 1
            if not head_name:      missing["Household Head Name"] += 1

            status_counts[admin_status] += 1
            record_type_counts[admin_status][record_type] += 1

            ws.append([
                ind_id,
                boundary.get("country", ""),
                boundary.get("state", ""),
                boundary.get("lga", ""),
                boundary.get("ward", ""),
                boundary.get("healthFacility", ""),
                doc.get("userName", ""),
                child_name,
                doc.get("age", ""),
                additional.get("gender", ""),
                beneficiary_id,
                head_name,
                doc.get("latitude", ""),
                doc.get("longitude", ""),
                doc.get("productName", ""),
                doc.get("taskDates", ""),
                additional.get("cycleIndex", ""),
                admin_status,
                doc.get("@timestamp", ""),
                record_type,
                quantity if admin_status == "ADMINISTRATION_SUCCESS" else "",
                quantity if admin_status == "VISITED" else "",
            ])
            total += 1

        if total % 100_000 == 0:
            print(f"  {total:,} rows written...")

    wb.save(output_path)
    print(f"\nReport saved to: {output_path} ({total:,} rows)")
    return total, status_counts, record_type_counts, missing


# ====================================================================
# MAIN
# ====================================================================

# Pass 1 — detect duplicates; collect unique individual IDs via intern table
dup_keys, ind_ids_set, pass1_total = pass1_count_records()

if pass1_total == 0:
    print("No data found for the given campaign and date range. Generating empty report.")
    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append(EXPORT_COLUMNS)
    output_path = os.path.join(os.getcwd(), f"{FILE_NAME}.xlsx")
    wb.save(output_path)
    print(f"Empty report saved to: {output_path}")
    sys.exit(0)

# Enrichment — build name lookup dicts (individual + HHM)
ind_ids_list = list(ind_ids_set)
del ind_ids_set  # free pointer set (~21 MB for 2.6 M IDs)
ind_id_to_name, ind_id_to_head_name = build_enrichment_maps(ind_ids_list)
del ind_ids_list

# Pass 2 — stream to Excel
output_path = os.path.join(os.getcwd(), f"{FILE_NAME}.xlsx")
total, status_counts, record_type_counts, missing = pass2_stream_to_excel(
    dup_keys, ind_id_to_name, ind_id_to_head_name, output_path
)


# ====================================================================
# SUMMARY
# ====================================================================
print("\n" + "=" * 70)
print("DATA QUALITY SUMMARY")
print("=" * 70)
for col, miss in missing.items():
    pct = 100 * miss / total if total > 0 else 0
    print(f"  {col:<30}: {miss:>7} missing ({pct:>6.2f}%)")

print("\n" + "=" * 70)
print("STATUS-WISE SUMMARY")
print("=" * 70)
print(f"\n{'Administration Status':<35} {'Records':>10} {'%':>8}")
print("-" * 55)
for status in sorted(status_counts):
    count = status_counts[status]
    pct   = 100 * count / total if total > 0 else 0
    print(f"  {status:<33} {count:>10} {pct:>7.2f}%")
print("-" * 55)
print(f"  {'GRAND TOTAL':<33} {total:>10} {'100.00%':>8}")

print("\n" + "=" * 70)
print("RECORD TYPE BREAKDOWN PER STATUS")
print("=" * 70)
for status in sorted(record_type_counts):
    t = status_counts[status]
    u = record_type_counts[status]["unique"]
    d = record_type_counts[status]["duplicate"]
    print(f"\n  {status}")
    print(f"    Total     : {t:>7}")
    print(f"    Unique    : {u:>7} ({100*u/t if t else 0:>6.2f}%)")
    print(f"    Duplicate : {d:>7} ({100*d/t if t else 0:>6.2f}%)")

unique_individuals = len(ind_id_to_name)
print("\n" + "=" * 70)
print(f"Unique Individuals      : {unique_individuals:,}")
print(f"Total Records Exported  : {total:,}")
if unique_individuals:
    print(f"Avg Records/Individual  : {total / unique_individuals:.2f}")
print("=" * 70 + "\n")
